import asyncio
import re

import config
import nest_asyncio
import os
import pytz
import threading
import json
import pandas as pd
import telegram.error
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ChatPermissions, \
    BotCommand, BotCommandScopeDefault, BotCommandScopeChat, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, ContextTypes
from datetime import datetime, timedelta
from flask import Flask

from aiocron import crontab
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

from apscheduler.schedulers.background import BackgroundScheduler


from telegram.ext import Application

nest_asyncio.apply()

global muted_users


scheduler = BackgroundScheduler(timezone="Europe/Kiev")

DATA_FILE = "data.json"
EXCEL_FILE = "user_data_export.xlsx"

application = None

app = Flask(__name__)

@app.route("/")
def index():
    return "@Supp0rtsBot"


def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)

with open(DATA_FILE, "r") as file:
    config = json.load(file)

def get_current_time_kiev():
    kiev_tz = pytz.timezone('Europe/Kiev')
    now = datetime.now(kiev_tz)
    return now.strftime("%H:%M; %d/%m/%Y")

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_data(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "users": [],
            "admins": [],
            "programmers": [],
            "bot_token": "",
            "owner_id": "",
            "chat_id": "",
            "total_score": 0.0,
            "num_of_ratings": 0,
            "sent_messages": {},
            "muted_users": {}
        }
    except json.JSONDecodeError:
        print("–ü–æ–º–∏–ª–∫–∞: –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–∏–π —Ñ–æ—Ä–º–∞—Ç JSON.")
        return {}

def load_sent_messages():
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    return data.get("sent_messages", {})

def save_sent_messages(sent_messages):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    data["sent_messages"] = sent_messages
    with open(DATA_FILE, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)

def load_muted_users_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    muted_users = {}
    for user in data.get("users", []):
        if user.get("mute", False):
            mute_end = user.get("mute_end")
            if mute_end:
                mute_end = datetime.strptime(mute_end, "%H:%M; %d/%m/%Y")
            muted_users[user["id"]] = {
                "first_name": user.get("first_name"),
                "username": user.get("username"),
                "expiration": mute_end,
                "reason": user.get("reason")
            }
    return muted_users

def load_users_info(json_file=DATA_FILE):
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get("users", [])
    except FileNotFoundError:
        print(f"–ü–æ–º–∏–ª–∫–∞: –§–∞–π–ª '{json_file}' –Ω–µ –∑–Ω–∞–π–¥–µ–Ω.")
        return []
    except json.JSONDecodeError:
        print("–ü–æ–º–∏–ª–∫–∞: –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω–∏–π —Ñ–æ—Ä–º–∞—Ç JSON.")
        return []

def load_chat_id_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    chat_id = data.get("chat_id")
    return chat_id

def load_bottocen_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    bot_token = data.get("bot_token")
    return bot_token

def update_data_json(data):
    with open(DATA_FILE, "w") as file:
        json.dump(data, file, indent=4, ensure_ascii=False)


users_info = load_users_info()
muted_users = load_muted_users_from_file()

CREATOR_CHAT_ID = load_chat_id_from_file()
BOTTOCEN = load_bottocen_from_file()



async def start(update: Update, context):
    user = update.message.from_user
    chat_id = update.effective_chat.id

    if chat_id == CREATOR_CHAT_ID:
        await update.message.reply_text("–ö–æ–º–∞–Ω–¥–∞ /start –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞ –≤ —Ü—ñ–π –≥—Ä—É–ø—ñ.")
        return

    user_found = False
    for u in config["users"]:
        if u["id"] == str(user.id):
            user_found = True
            break

    if not user_found:
        new_user = {
            "id": str(user.id),
            "username": user.username or "–ù–µ –≤–∫–∞–∑–∞–Ω–æ",
            "first_name": user.first_name or "–ù–µ –≤–∫–∞–∑–∞–Ω–æ",
            "join_date": get_current_time_kiev(),
            "rating": 0,
            "mute": False,
            "mute_end": None,
            "reason": None
        }
        config["users"].append(new_user)
        save_data(config)

    keyboard = [
        ["/start", "/rate"],
        ["/message", "/stopmessage"],
        ["/fromus", "/help"],
    ]

    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "–ü—Ä–∏–≤—ñ—Ç! –Ø –≤–∞—à –±–æ—Ç –ø—ñ–¥—Ç—Ä–∏–º–∫–∏. –í–≤–µ–¥—ñ—Ç—å –∫–æ–º–∞–Ω–¥—É /rate –¥–ª—è –æ—ñ–Ω–∫–∏ –±–æ—Ç–∞, /message –¥–ª—è –Ω–∞–ø–∏—Å–∞–Ω–∏—è –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞ –∞–±–æ /help –¥–ª—è –æ—Ç—Ä–∏–º–∞–Ω–Ω—è —ñ–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—ó –ø—Ä–æ –∫–æ–º–∞–Ω–¥–∏.",
        reply_markup=reply_markup
    )

async def rate(update: Update, context):
    user_id = update.message.from_user.id

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_rating = None
    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            user_rating = user['rating']
            break

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    rating_text = f"–ó–∞–≥–∞–ª—å–Ω–∞ –æ—Ü—ñ–Ω–∫–∞: {round(average_rating, 1)}‚≠êÔ∏è\n–í–∞—à –ø–æ–ø–µ—Ä–µ–¥–Ω—ñ–π –≤—ñ–¥–≥—É–∫: {user_rating}‚≠êÔ∏è"

    keyboard = [
        [InlineKeyboardButton("0.5‚≠êÔ∏è", callback_data='0.5'), InlineKeyboardButton("1‚≠êÔ∏è", callback_data='1')],
        [InlineKeyboardButton("1.5‚≠êÔ∏è", callback_data='1.5'), InlineKeyboardButton("2‚≠êÔ∏è", callback_data='2')],
        [InlineKeyboardButton("2.5‚≠êÔ∏è", callback_data='2.5'), InlineKeyboardButton("3‚≠êÔ∏è", callback_data='3')],
        [InlineKeyboardButton("3.5‚≠êÔ∏è", callback_data='3.5'), InlineKeyboardButton("4‚≠êÔ∏è", callback_data='4')],
        [InlineKeyboardButton("4.5‚≠êÔ∏è", callback_data='4.5'), InlineKeyboardButton("5‚≠êÔ∏è", callback_data='5')],
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(f"{rating_text}\n–û–±–µ—Ä—ñ—Ç—å –æ—Ü—ñ–Ω–∫—É:", reply_markup=reply_markup)

async def button_callback(update: Update, context):
    query = update.callback_query
    user_id = query.from_user.id
    new_rating = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_found = False
    previous_rating = 0

    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            previous_rating = user.get('rating', 0)
            user['rating'] = new_rating
            user_found = True
            break

    if not user_found:
        new_user = {
            'id': str(user_id),
            'first_name': query.from_user.first_name,
            'username': query.from_user.username,
            'join_date': datetime.now().strftime("%H:%M; %d/%m/%Y"),
            'rating': new_rating,
            'mute': False,
            'mute_end': None,
            'reason': None
        }
        data['users'].append(new_user)

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    if previous_rating == 0:
        num_of_ratings += 1
        total_score += new_rating
    else:
        total_score = total_score - previous_rating + new_rating

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    await query.edit_message_text(
        f"–î—è–∫—É—î–º–æ –∑–∞ –≤–∞—à –≤—ñ–¥–≥—É–∫! –í–∞—à–∞ –æ—Ü—ñ–Ω–∫–∞: {new_rating}‚≠êÔ∏è\n–ó–∞–≥–∞–ª—å–Ω–∞ –æ—Ü—ñ–Ω–∫–∞: {round(average_rating, 1)}‚≠êÔ∏è"
    )

async def button(update: Update, context):
    global total_score, num_of_ratings

    query = update.callback_query
    await query.answer()

    selected_rate = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    total_score = data.get("total_score", 0) + selected_rate
    num_of_ratings = data.get("num_of_ratings", 0) + 1

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings

    user_id = query.from_user.id
    if user_id in users_info:
        users_info[user_id]['rating'] = selected_rate

    await query.edit_message_text(
        f"–î—è–∫—É—î–º–æ –∑–∞ –≤–∞—à –≤—ñ–¥–≥—É–∫! –í–∞—à–∞ –æ—Ü—ñ–Ω–∫–∞: {selected_rate}‚≠êÔ∏è\n–ó–∞–≥–∞–ª—å–Ω–∞ –æ—Ü—ñ–Ω–∫–∞: {round(average_rating, 1)}‚≠êÔ∏è")

async def auto_delete_message(bot, chat_id, message_id, delay):
    await asyncio.sleep(delay)
    await bot.delete_message(chat_id=chat_id, message_id=message_id)

async def message(update: Update, context):
    user_id = update.message.from_user.id
    muted_users = load_muted_users_from_file()


    if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
        reply = await update.message.reply_text("–í–∏ –≤ –º—É—Ç—ñ –π –Ω–µ –º–æ–∂–µ—Ç–µ –Ω–∞–¥—Å–∏–ª–∞—Ç–∏ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è.")
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
        return

    reply = await update.message.reply_text(
        "–í–≤–µ–¥—ñ—Ç—å –≤–∞—à–µ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è, —ñ –π–æ–≥–æ –±—É–¥–µ –≤—ñ–¥–ø—Ä–∞–≤–ª–µ–Ω–æ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞. –í–≤–µ–¥—ñ—Ç—å /stopmessage, —â–æ–± –∑–∞–≤–µ—Ä—à–∏—Ç–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å."
    )

    context.user_data['waiting_for_message'] = True

    await asyncio.create_task(
        auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))

async def stopmessage(update: Update, context):
    if context.user_data.get('waiting_for_message'):
        reply = await update.message.reply_text("–í–∏ –∑–∞–≤–µ—Ä—à–∏–ª–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å.")
        context.user_data['waiting_for_message'] = False
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    else:
        await update.message.reply_text("–í–∏ –Ω–µ –≤ —Ä–µ–∂–∏–º—ñ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å.")

async def help(update: Update, context):
    if str(update.message.chat.id) == str(CREATOR_CHAT_ID):
        help_text = (
            "–î–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏ –≤ –≥—Ä—É–ø—ñ:\n"
            "–í—ñ–¥–ø–æ–≤—ñ—Å—Ç–∏ –Ω–∞ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –±–æ—Ç–∞ - –ù–∞–¥—ñ—Å–ª–∞—Ç–∏ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—É, —è–∫–∏–π –Ω–∞–¥—ñ—Å–ª–∞–≤ —Ü–µ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è.\n"
            "/mute <—á–∞—Å> <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> '–ø—Ä–∏—á–∏–Ω–∞' - –ó–∞–º—É—Ç–∏—Ç–∏ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞ –Ω–∞ –≤–∫–∞–∑–∞–Ω–∏–π —á–∞—Å.\n"
            "/unmute <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> - –†–æ–∑–º—É—Ç–∏—Ç–∏ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞.\n"
            "/mutelist - –ü–æ–∫–∞–∑–∞—Ç–∏ —Å–ø–∏—Å–æ–∫ –∑–∞–º—É—á–µ–Ω–∏—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤.\n"
            "/alllist - –ü–æ–∫–∞–∑–∞—Ç–∏ –≤—Å—ñ—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤.\n"
            "/allmessage <–ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è> - –ù–∞–¥—ñ—Å–ª–∞—Ç–∏ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –≤—Å—ñ–º –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º.\n"
            "/fromus - –Ü–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—è –ø—Ä–æ —Å—Ç–≤–æ—Ä—é–≤–∞—á–∞.\n"
            "/help - –ü–æ–∫–∞–∑–∞—Ç–∏ –¥–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏.\n"
            "/info - –ü–æ–∫–∞–∑–∞—Ç–∏ —ñ–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—é –ø—Ä–æ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç—ñ–≤ —Ç–∞ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä—ñ–≤.\n"
            "/admin <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> - –î–æ–¥–∞—Ç–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞.\n"
            "/deleteadmin <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> - –í–∏–¥–∞–ª–∏—Ç–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞.\n"
            "/programier <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> - –î–æ–¥–∞—Ç–∏ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç–∞.\n"
            "/deleteprogramier <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á> - –í–∏–¥–∞–ª–∏—Ç–∏ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç–∞.\n"
            "/get_alllist - –û—Ç—Ä–∏–º–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏.\n"
            "/set_alllist - –ó–∞–ø–∏—Å–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏.\n"
        )
    elif str(update.message.chat.id) == str(-1002358066044):
        help_text = (
            "–î–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏ –≤ –≥—Ä—É–ø—ñ:\n"
            "/get_alllist - –û—Ç—Ä–∏–º–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏.\n"
            "/set_alllist - –ó–∞–ø–∏—Å–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏.\n"
        )
    else:
        help_text = (
            "–î–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏ –≤ –±–æ—Ç—ñ:\n"
            "/start - –ó–∞–ø—É—Å—Ç–∏—Ç–∏ –±–æ—Ç–∞.\n"
            "/rate - –ó–∞–ª–∏—à–∏—Ç–∏ –≤—ñ–¥–≥—É–∫.\n"
            "/message - –ü–æ—á–∞—Ç–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.\n"
            "/stopmessage - –ó–∞–≤–µ—Ä—à–∏—Ç–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å.\n"
            "/fromus - –Ü–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—è –ø—Ä–æ —Å—Ç–≤–æ—Ä—é–≤–∞—á–∞.\n"
            "/help - –ü–æ–∫–∞–∑–∞—Ç–∏ –¥–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏.\n"
        )

    await update.message.reply_text(help_text)

async def fromus(update: Update, context):
    await update.message.reply_text(
        "*Skeleton*  –ù–∞–ø–∏—Å–≤ –±–æ—Ç–∞\n–ü–æ—Ä—Ç—Ñ–æ–ª—ñ–æ:  ```https://www.linkedin.com/in/artem-k-972a41344/``` \n –¢–µ–ª–µ–≥—Ä–∞–º –∫–∞–Ω–∞–ª –∑ —É—Å—ñ–º–∞ –ø—Ä–æ—î–∫—Ç–∞–º–∏: ```https://t.me/AboutMyProjects```\n –ü–æ –≤—Å—ñ–º –ø–∏—Ç–∞–Ω–Ω—è–º –ø–∏—à—ñ—Ç—å –≤ —Ü—å–æ–≥–æ –±–æ—Ç–∞",
        parse_mode="MarkdownV2"
    )

async def info(update: Update, context: CallbackContext):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    programmers = data.get("programmers", [])
    admins = data.get("admins", [])

    programmer_list = "\n".join(programmers) if programmers else "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–≥—Ä–∞–º–º–∏—Å—Ç–æ–≤ –ø—É—Å—Ç."
    admin_list = "\n".join(admins) if admins else "–°–ø–∏—Å–æ–∫ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤ –ø—É—Å—Ç."

    await update.message.reply_text(f"–ü—Ä–æ–≥—Ä–∞–º–º—ñ—Å—Ç–∏:\n{programmer_list}\n\n–ê–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∏:\n{admin_list}")




async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sent_messages = load_sent_messages()
    muted_users = load_muted_users_from_file()
    if context.user_data.get("awaiting_file"):
        if update.message.document:
            document = update.message.document
            file_path = "uploaded_file.xlsx"

            file = await document.get_file()
            await file.download_to_drive(file_path)

            try:
                wb = load_workbook(file_path)

                sheet_all_user = wb["AllUser"]
                sheet_admins = wb["Admins"]
                sheet_programmers = wb["Programmers"]
                sheet_general_info = wb["GeneralInfo"]
                sheet_sent_messages = wb["SentMessages"]

                updated_users = []
                muted_users = {}
                sent_messages = {}

                for row in sheet_all_user.iter_rows(min_row=2, values_only=True):
                    if len(row) < 8:
                        continue

                    user_data = {
                        "id": str(row[0]),
                        "first_name": row[1],
                        "username": row[2],
                        "join_date": row[3].strftime("%H:%M; %d/%m/%Y") if isinstance(row[3], datetime) else str(
                            row[3]),
                        "rating": int(row[4]) if row[4] is not None else 0,
                        "mute": bool(row[5]),
                        "mute_end": row[6].strftime("%H:%M; %d/%m/%Y") if isinstance(row[6], datetime) else str(row[6]),
                        "reason": row[7]
                    }
                    updated_users.append(user_data)

                    if user_data["mute"]:
                        muted_users[user_data["username"]] = True

                for row in sheet_sent_messages.iter_rows(min_row=2, values_only=True):
                    if len(row) < 2 or not row[0] or not row[1]:
                        continue
                    sent_messages[str(row[0])] = row[1]

                admins = [row[0] for row in sheet_admins.iter_rows(min_row=2, values_only=True)]
                programmers = [row[0] for row in sheet_programmers.iter_rows(min_row=2, values_only=True)]

                bot_token = sheet_general_info.cell(row=2, column=1).value or ""
                owner_id = sheet_general_info.cell(row=2, column=2).value or ""
                chat_id = sheet_general_info.cell(row=2, column=3).value or ""
                total_score = float(sheet_general_info.cell(row=2, column=4).value or 0)
                num_of_ratings = int(sheet_general_info.cell(row=2, column=5).value or 0)

                data = {
                    "users": updated_users,
                    "muted_users": muted_users,
                    "admins": admins,
                    "programmers": programmers,
                    "bot_token": bot_token,
                    "owner_id": owner_id,
                    "chat_id": chat_id,
                    "total_score": total_score,
                    "num_of_ratings": num_of_ratings,
                    "sent_messages": sent_messages,
                }

                with open(DATA_FILE, "w", encoding="utf-8") as json_file_obj:
                    json.dump(data, json_file_obj, ensure_ascii=False, indent=4,
                              default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj,
                                                                                                datetime) else None)

                await update.message.reply_text("–§–∞–π–ª —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω!")

            except Exception as e:
                await update.message.reply_text(f"–ü–æ–º–∏–ª–∫–∞ –ø—Ä–∏ –æ–±—Ä–æ–±—Ü—ñ —Ñ–∞–π–ª–∞: {e}")

            finally:
                context.user_data["awaiting_file"] = False
        else:
            await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ Excel-—Ñ–∞–π–ª.")
    elif (str(update.message.chat.id)) != (str(CREATOR_CHAT_ID)):
        user_id = update.message.from_user.id
        if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
            reply = await update.message.reply_text("–í–∏ –≤ –º—É—Ç—ñ –π –Ω–µ –º–æ–∂–µ—Ç–µ –Ω–∞–¥—Å–∏–ª–∞—Ç–∏ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

        if context.user_data.get('waiting_for_message'):
            user_name = update.effective_user.first_name
            user_username = update.effective_user.username if update.effective_user.username else "–Ω–µ–º–∞—î —ñ–º–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞"
            current_time = get_current_time_kiev()
            user_message = ""
            first_message = f'–ü–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –≤—ñ–¥ **{user_name}**; ```@{user_username}``` ```{user_id}``` \n{current_time}:'

            if update.message.text:
                user_message = update.message.text
                first_message += f'\n{user_message}'

            message_to_admin = await context.bot.send_message(chat_id=CREATOR_CHAT_ID, text=first_message, parse_mode="MarkdownV2")
            sent_messages[message_to_admin.message_id] = update.effective_user.id

            if update.message.photo:
                photo_file_id = update.message.photo[-1].file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_photo(chat_id=CREATOR_CHAT_ID, photo=photo_file_id, caption=caption, reply_to_message_id=message_to_admin.message_id)

            elif update.message.document:
                document_file_id = update.message.document.file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_document(chat_id=CREATOR_CHAT_ID, document=document_file_id, caption=caption, reply_to_message_id=message_to_admin.message_id)
            elif update.message.sticker:
                sticker_file_id = update.message.sticker.file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_sticker(chat_id=CREATOR_CHAT_ID, sticker=sticker_file_id, reply_to_message_id=message_to_admin.message_id)
            elif update.message.voice:
                voice_file_id = update.message.voice.file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_voice(chat_id=CREATOR_CHAT_ID, voice=voice_file_id, caption=caption, reply_to_message_id=message_to_admin.message_id)
            elif update.message.video:
                video_file_id = update.message.video.file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_video(chat_id=CREATOR_CHAT_ID, video=video_file_id, caption=caption, reply_to_message_id=message_to_admin.message_id)
            elif update.message.video_note:
                video_note_file_id = update.message.video_note.file_id
                caption = update.message.caption if update.message.caption else ''
                await context.bot.send_video_note(chat_id=CREATOR_CHAT_ID, video_note=video_note_file_id, reply_to_message_id=message_to_admin.message_id)

            reply = await update.message.reply_text("–í–∞—à–µ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –Ω–∞–¥—ñ—Å–ª–∞–Ω–æ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
            sent_messages[update.message.message_id] = update.message.from_user.id
            save_sent_messages(sent_messages)
        else:
            await update.message.reply_text("–í–≤–µ–¥—ñ—Ç—å /message, —â–æ–± –Ω–∞–¥—Å–∏–ª–∞—Ç–∏ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞.")
    else:
        if update.effective_user.id != context.bot.id:
            if update.message.reply_to_message:
                if update.message.reply_to_message.from_user.id == context.bot.id:
                    original_message_id = str(update.message.reply_to_message.message_id)
                    if original_message_id in sent_messages:
                        original_user_id = sent_messages[original_message_id]
                        reply_text = update.message.text if update.message.text else ""
                        for user in config['users']:
                            if str(user['id']) == str(original_user_id):
                                user_name = user['first_name']
                                break

                        if update.message.photo:
                            photo_file_id = update.message.photo[-1].file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_photo(chat_id=original_user_id, photo=photo_file_id, caption=caption)

                        elif update.message.document:
                            document_file_id = update.message.document.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_document(chat_id=original_user_id, document=document_file_id, caption=caption)
                        elif update.message.sticker:
                            sticker_file_id = update.message.sticker.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_sticker(chat_id=original_user_id, sticker=sticker_file_id)

                        elif update.message.voice:
                            voice_file_id = update.message.voice.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_voice(chat_id=original_user_id, voice=voice_file_id, caption=caption)

                        elif update.message.video:
                            video_file_id = update.message.video.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video(chat_id=original_user_id, video=video_file_id, caption=caption)

                        elif update.message.video_note:
                            video_note_file_id = update.message.video_note.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video_note(chat_id=original_user_id, video_note=video_note_file_id)
                        else:
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_message(chat_id=original_user_id, text=reply_text)
                        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á—É { user_name } –±—É–ª–æ –Ω–∞–¥—ñ—Å–ª–∞–Ω–æ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è")
                        sent_messages[update.message.message_id] = update.message.from_user.id
                        save_sent_messages(sent_messages)


async def mute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    message_text = update.message.text.split()

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")
        return

    mute_time = 300
    reason = "–ü–æ —Ä—ñ—à–µ–Ω–Ω—é –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞"
    username = None

    if len(context.args) > 0:
        if context.args[0].isdigit():
            mute_time = int(context.args[0])
            username = context.args[1].lstrip('@') if len(context.args) > 1 else None
        else:
            username = context.args[0].lstrip('@')

    reason_match = re.search(r'["\'](.*?)["\']', update.message.text)
    if reason_match:
        reason = reason_match.group(1)

    if not username:
        await update.message.reply_text("–ù–µ –≤–∫–∞–∑–∞–Ω–æ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞ –¥–ª—è –º—É—Ç–∞.")
        return

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username),
                None)

    if not user:
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {username} –Ω–µ –∑–Ω–∞–π–¥–µ–Ω.")
        return

    if user["id"] == config["owner_id"]:
        await update.message.reply_text("–ù–µ–º–æ–∂–ª–∏–≤–æ –∑–∞–º—É—Ç–∏—Ç–∏ –≤–ª–∞—Å–Ω–∏–∫–∞ —á–∞—Ç—É.")
        return

    if user["mute"]:
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {user['first_name']} –≤–∂–µ –±—É–≤ –∑–∞–º—É—á–µ–Ω–∏–π.")


    user["mute"] = True
    user["mute_end"] = (datetime.now() + timedelta(seconds=mute_time)).strftime("%H:%M; %d/%m/%Y")
    user["reason"] = reason

    config["muted_users"][username] = True
    save_data(config)

    mute_permissions = ChatPermissions(can_send_messages=False)
    await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"], permissions=mute_permissions)
    await context.bot.send_message(chat_id=user["id"],
                                   text=f"–í–∞—Å –∑–∞–º—É—Ç–∏–ª–∏ –Ω–∞ {str(timedelta(seconds=mute_time))}\n–ü—Ä–∏—á–∏–Ω–∞: {reason}")
    await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{user['username']} –∑–∞–º—É—á–µ–Ω–∏–π.")

async def unmute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("–í–∏–∫–æ—Ä–∏—Å—Ç–æ–≤—É–π—Ç–µ: /unmute <–∫–æ—Ä–∏—Å—Ç—É–≤–∞—á>")
        return



    username = context.args[0].lstrip('@')

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username), None)

    if user and user["mute"]:
        user["mute"] = False
        user["mute_end"] = None
        user["reason"] = None

        config["muted_users"].pop(username, None)
        save_data(config)

        mute_permissions = ChatPermissions(can_send_messages=True)
        await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"], permissions=mute_permissions)
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{user['username']} –±—É–≤ —Ä–æ–∑–º—É—á–µ–Ω–∏–π.")
    else:
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {username} –Ω–µ –∑–Ω–∞–π–¥–µ–Ω –∞–±–æ –Ω–µ –±—É–≤ –∑–∞–º—É—á–µ–Ω–∏–π.")

async def admin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç–∞–º.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("–í–∏–∫–æ—Ä–∏—Å—Ç–æ–≤—É–π—Ç–µ: /admin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{username} –≤–∂–µ —î –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º.")
    else:
        config["admins"].append(username)
        save_data(config)
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{username} –¥–æ–¥–∞–Ω –≤ —Å–ø–∏—Å–æ–∫ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä—ñ–≤.")

async def deleteadmin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç–∞–º.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("–í–∏–∫–æ—Ä–∏—Å—Ç–æ–≤—É–π—Ç–µ: /deleteadmin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        config["admins"].remove(username)
        save_data(config)
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{username} –≤–∏–¥–∞–ª–µ–Ω –∑—ñ —Å–ø–∏—Å–∫—É –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä—ñ–≤.")
    else:
        await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á @{username} –Ω–µ –∑–Ω–∞–π–¥–µ–Ω.")

async def programier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            new_programmer = context.args[0].replace("@", "")
            if new_programmer not in config["programmers"]:
                config["programmers"].append(new_programmer)
                save_data(config)
                await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {new_programmer} –¥–æ–¥–∞–Ω –≤ —Å–ø–∏—Å–æ–∫ –ø—Ä–æ–≥—Ä–∞–º–º—ñ—Å—Ç—ñ–≤.")
            else:
                await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {new_programmer} –≤–∂–µ —î –≤ —Å–ø–∏—Å–∫—É –ø—Ä–æ–≥—Ä–∞–º–º–∏—Å—Ç—ñ–≤.")
        else:
            await update.message.reply_text("–í–∏–∫–æ—Ä–∏—Å—Ç–æ–≤—É–π—Ç–µ: /programier @username")
    else:
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ –ª–∏—à–µ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")

async def deleteprogramier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            removed_programmer = context.args[0].replace("@", "")
            if removed_programmer == "ArtemKirss":
                await update.message.reply_text(f"–ù–µ–º–æ–∂–ª–∏–≤–æ –≤–∏–¥–∞–ª–∏—Ç–∏ {removed_programmer} –∑—ñ —Å–ø–∏—Å–∫—É –ø—Ä–æ–≥—Ä–∞–º–º–∏—Å—Ç–æ–≤.")
            elif removed_programmer in config["programmers"]:
                config["programmers"].remove(removed_programmer)
                save_data(config)
                await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {removed_programmer} –≤–∏–¥–∞–ª–µ–Ω –∑—ñ —Å–ø–∏—Å–∫—É –ø—Ä–æ–≥—Ä–∞–º–º–∏—Å—Ç—ñ–≤.")
            else:
                await update.message.reply_text(f"–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á {removed_programmer} –Ω–µ —î –ø—Ä–æ–≥—Ä–∞–º–º–∏—Å—Ç–æ–º.")
        else:
            await update.message.reply_text("–í–∏–∫–æ—Ä–∏—Å—Ç–æ–≤—É–π—Ç–µ: /deleteprogramier @username")
    else:
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ –ª–∏—à–µ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")

async def mutelist(update: Update, context):
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "–ó–∞–º—É—á–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info.get('mute_end', '–ù–µ–≤—ñ–¥–æ–º–æ')
            reason = mute_info.get('reason', '–ë–µ–∑ –ø—Ä–∏—á–∏–Ω–∏')

            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=int(user_id))
            user_fullname = user_info.user.first_name or "–ù–µ–≤—ñ–¥–æ–º–∏–π"
            username = user_info.user.username or "–ù–µ–º–∞—î —ñ–º–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞"

            join_date = mute_info.get('join_date', '–ù–µ–≤—ñ–¥–æ–º–∞')
            rating = mute_info.get('rating', 0)
            mute_symbol = "üîá"

            admins_sumdol = "üë®üèª‚Äçüíº"
            if username in admins:
                admins_sumdol = "üëÆüèª‚Äç‚ôÇÔ∏è"
            if username in programmers:
                admins_sumdol = "üë®üèª‚Äçüíª"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"–ó–∞–ª–∏—à–∏–ª–æ—Å—å: {expiration}\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: {reason}\n"
                f"–î–∞—Ç–∞ –∑–∞—Ö–æ–¥—É: {join_date}\n"
                f"–û—Ü—ñ–Ω–∫–∞: {rating}‚≠êÔ∏è\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "–ù–µ–º–∞—î –∑–∞–º—É—á–µ–Ω–∏—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)

async def alllist(update: Update, context: CallbackContext):
    global mute_symbol
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ –ª–∏—à–µ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    users_info = {user['id']: user for user in data.get("users", [])}
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "–ö–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ:\n"
    unique_users = {user['id'] for user in data.get("users", [])}

    if unique_users:
        for user_id in unique_users:
            user_data = users_info.get(str(user_id), {})
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "–ù–µ–≤—ñ–¥–æ–º–∏–π"
            username = user_info.user.username or "–ù–µ–º–∞—î —ñ–º–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞"
            join_date = user_data.get('join_date', '–ù–µ–≤—ñ–¥–æ–º–∞')
            rating = user_data.get('rating', 0)

            admins_sumdol = "üë®üèª‚Äçüíº"
            if username in admins:
                admins_sumdol = "üëÆüèª‚Äç‚ôÇÔ∏è"
            if username in programmers:
                admins_sumdol = "üë®üèª‚Äçüíª"

            mute_symbol = "üîá" if str(user_id) in muted_users else "üîä"

            response += f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n–î–∞—Ç–∞ –∑–∞—Ö–æ–¥—É: {join_date}\n–û—Ü—ñ–Ω–∫–∞: {rating}‚≠êÔ∏è\n"
            response += "-------------------------------------------------------------------------\n"
    else:
        response += "–ù–µ–º–∞—î –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤.\n"
        response += "-------------------------------------------------------------------------\n"

    response += "==========================================\n"
    response += "\n"
    response += "==========================================\n"
    response += "–ó–∞–º—É—á–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info['mute_end'] or "–ù–µ–≤—ñ–¥–æ–º–æ"
            reason = mute_info.get('reason', "–ë–µ–∑ –ø—Ä–∏—á–∏–Ω–∏")
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "–ù–µ–≤—ñ–¥–æ–º–∏–π"
            username = user_info.user.username or "–ù–µ–º–∞—î —ñ–º–µ–Ω—ñ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞"
            user_data = users_info.get(str(user_id), {})
            join_date = user_data.get('join_date', '–ù–µ–≤—ñ–¥–æ–º–∞')
            rating = user_data.get('rating', 0)

            admins_sumdol = "üë®üèª‚Äçüíº"
            if username in admins:
                admins_sumdol = "üëÆüèª‚Äç‚ôÇÔ∏è"
            if username in programmers:
                admins_sumdol = "üë®üèª‚Äçüíª"

            mute_symbol = "üîá"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"–ó–∞–ª–∏—à–∏–ª–æ—Å—å: {str(expiration).split('.')[0]}\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: {reason}\n"
                f"–î–∞—Ç–∞ –∑–∞—Ö–æ–¥—É: {join_date}\n"
                f"–û—Ü—ñ–Ω–∫–∞: {rating}‚≠êÔ∏è\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "–ù–µ–º–∞—î –∑–∞–º—É—á–µ–Ω–∏—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)

async def allmessage(update: Update, context):
    user = update.message.from_user.username

    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º –±–æ—Ç–∞.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    if not context.args:
        await update.message.reply_text("–ë—É–¥—å –ª–∞—Å–∫–∞, —É–∫–∞–∂—ñ—Ç—å —Ç–µ–∫—Å—Ç –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –ø—ñ—Å–ª—è –∫–æ–º–∞–Ω–¥–∏.")
        return

    message_text = update.message.text.split(' ', 1)[1]

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        config = json.load(file)

    users = config.get("users", [])

    for user_data in users:
        user_id = user_data.get("id")
        if user_id:
            try:
                await context.bot.send_message(chat_id=user_id, text=message_text)
            except Exception as e:
                print(f"–ü–æ–º–∏–ª–∫–∞ –ø—Ä–∏ –≤—ñ–¥–ø—Ä–∞–≤—Ü—ñ –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—É {user_id}: {e}")

    await update.message.reply_text("–ü–æ–≤—ñ–¥–æ–º–ª–µ–Ω–Ω—è –≤—ñ–¥–ø—Ä–∞–≤–ª–µ–Ω–æ –≤—Å—ñ–º –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º.")

def is_programmer(username):
    return username in config["programmers"]

def is_admin(username):
    return username in config["admins"]


async def get_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")
        return
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])

        users_df = all_users_df[all_users_df["mute"] == False]
        muted_df = all_users_df[all_users_df["mute"] == True]

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rtsBot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
            username_cell = row[2]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value), False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:8]:
                cell.fill = fill_color

        workbook.save(excel_file)

        await update.message.reply_document(document=open(excel_file, "rb"))

    except Exception as e:
        await update.message.reply_text(f"Error: {e}")

async def set_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("–¶—è –∫–æ–º–∞–Ω–¥–∞ –¥–æ—Å—Ç—É–ø–Ω–∞ —Ç—ñ–ª—å–∫–∏ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º.")
        return
    await update.message.reply_text("–ë—É–¥—å –ª–∞—Å–∫–∞ –ø—Ä–∏—à–ª—ñ—Ç—å Excel file –∑ –¥–∞–Ω–Ω–∏–º–∏.")
    context.user_data["awaiting_file"] = True

async def set_default_commands(application):
    commands = [
        BotCommand("start", "–ó–∞–ø—É—Å—Ç–∏—Ç–∏ –±–æ—Ç–∞"),
        BotCommand("rate", "–ó–∞–ª–∏—à–∏—Ç–∏ –≤—ñ–¥–≥—É–∫"),
        BotCommand("message", "–ü–æ—á–∞—Ç–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä–∞–º"),
        BotCommand("stopmessage", "–ó–∞–≤–µ—Ä—à–∏—Ç–∏ –≤–≤–µ–¥–µ–Ω–Ω—è –ø–æ–≤—ñ–¥–æ–º–ª–µ–Ω—å"),
        BotCommand("fromus", "–Ü–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—è –ø—Ä–æ —Å—Ç–≤–æ—Ä—é–≤–∞—á–∞"),
        BotCommand("help", "–ü–æ–∫–∞–∑–∞—Ç–∏ –¥–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeDefault())

async def set_creator_commands(application):
    commands = [
        BotCommand("mutelist", "–ü–æ–∫–∞–∑–∞—Ç–∏ —Å–ø–∏—Å–æ–∫ –∑–∞–º—É—á–µ–Ω–∏—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤"),
        BotCommand("alllist", "–ü–æ–∫–∞–∑–∞—Ç–∏ –≤—Å—ñ—Ö –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á—ñ–≤"),
        BotCommand("fromus", "–Ü–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—è –ø—Ä–æ —Å—Ç–≤–æ—Ä—é–≤–∞—á–∞"),
        BotCommand("help", "–ü–æ–∫–∞–∑–∞—Ç–∏ –¥–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏"),
        BotCommand("info", "–ü–æ–∫–∞–∑–∞—Ç–∏ —ñ–Ω—Ñ–æ—Ä–º–∞—Ü—ñ—é –ø—Ä–æ –ø—Ä–æ–≥—Ä–∞–º—ñ—Å—Ç—ñ–≤ —Ç–∞ –∞–¥–º—ñ–Ω—ñ—Å—Ç—Ä–∞—Ç–æ—Ä—ñ–≤"),
        BotCommand("get_alllist", "–û—Ç—Ä–∏–º–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏"),
        BotCommand("set_alllist", "–ó–∞–ø–∏—Å–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CREATOR_CHAT_ID))

async def set_save_commands(application):
    commands = [
        BotCommand("get_alllist", "–û—Ç—Ä–∏–º–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏"),
        BotCommand("set_alllist", "–ó–∞–ø–∏—Å–∞—Ç–∏ Exel —Ñ–∞–π–ª –∑ –∫–æ—Ä–∏—Å—Ç—É–≤–∞—á–∞–º–∏"),
        BotCommand("help", "–ü–æ–∫–∞–∑–∞—Ç–∏ –¥–æ—Å—Ç—É–ø–Ω—ñ –∫–æ–º–∞–Ω–¥–∏"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=-1002310142084))

async def send_user_list():
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])
        users_df = all_users_df[all_users_df["mute"] == False]
        muted_df = all_users_df[all_users_df["mute"] == True]

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rtsBot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
            username_cell = row[2]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value), False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:8]:
                cell.fill = fill_color

        workbook.save(excel_file)

        bot = Bot(token=BOTTOCEN)
        await bot.send_document(chat_id=-1002358066044, document=open(excel_file, "rb"))

    except Exception as e:
        bot = Bot(token=BOTTOCEN)
        await bot.send_message(chat_id=-1002358066044, text=f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –æ—Ç—á–µ—Ç–∞: {e}")

async def main():
    application = Application.builder().token(BOTTOCEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("rate", rate))
    application.add_handler(CommandHandler("message", message))
    application.add_handler(CommandHandler("stopmessage", stopmessage))
    application.add_handler(CommandHandler("fromus", fromus))
    application.add_handler(CommandHandler("help", help))
    application.add_handler(CommandHandler("mute", mute))
    application.add_handler(CommandHandler("unmute", unmute))
    application.add_handler(CommandHandler("mutelist", mutelist))
    application.add_handler(CommandHandler("alllist", alllist))
    application.add_handler(CommandHandler("allmessage", allmessage))
    application.add_handler(CommandHandler("admin", admin))
    application.add_handler(CommandHandler("deleteadmin", deleteadmin))
    application.add_handler(CommandHandler("programier", programier))
    application.add_handler(CommandHandler("deleteprogramier", deleteprogramier))
    application.add_handler(CommandHandler("info", info))
    application.add_handler(CommandHandler("get_alllist", get_alllist))
    application.add_handler(CommandHandler("set_alllist", set_alllist))

    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.ALL, handle_message))

    #await set_default_commands(application)
    #await set_creator_commands(application)
    #await set_save_commands(application)

    scheduler = AsyncIOScheduler(timezone=pytz.timezone("Europe/Kyiv"))
    scheduler.add_job(send_user_list, "cron", hour=0, minute=0)
    scheduler.start()

    application.run_polling()


if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()
    update_data_json(config)
    asyncio.run(main())