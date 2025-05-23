import asyncio
import re
import os
import nest_asyncio
import pytz
import threading
import json
import pandas as pd
import telegram.error
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ChatPermissions, \
    BotCommand, BotCommandScopeDefault, BotCommandScopeChat, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, \
    ContextTypes
from datetime import datetime, timedelta
from flask import Flask
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from apscheduler.schedulers.background import BackgroundScheduler

nest_asyncio.apply()

import logging
logging.basicConfig(
    filename='bot_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# ФУНКЦІЇ ДЛЯ РОБОТИ З JSON
def safe_json_read(file_path):
    """Безпечне читання JSON файлу з даними"""
    default_data = {
        "users": [],
        "muted_users": {},
        "banned_users": {},
        "admins": [],
        "programmers": [],
        "bot_token": "",
        "owner_id": "",
        "chat_id": "",
        "total_score": 0.0,
        "num_of_ratings": 0,
        "sent_messages": {},
        "topics": {},
        "user_topics": {}
    }

    if not os.path.exists(file_path):
        safe_json_write(default_data, file_path)
        return default_data

    try:
        for encoding in ['utf-8-sig', 'utf-8', 'cp1251']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    data = json.load(f)
                    for key in default_data.keys():
                        if key not in data:
                            data[key] = default_data[key]
                    return data
            except UnicodeDecodeError:
                continue
            except json.JSONDecodeError as je:
                print(f"Invalid JSON in {file_path}: {je}")
                continue

        print("All encoding attempts failed, creating new file")
        safe_json_write(default_data, file_path)
        return default_data
    except Exception as e:
        print(f"Critical read error: {e}")
        safe_json_write(default_data, file_path)
        return default_data

def safe_json_write(data, file_path):
    """Безпечний запис даних у JSON файл"""
    temp_file = file_path + '.tmp'
    try:
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        with open(temp_file, 'r', encoding='utf-8') as f:
            json.load(f)

        if os.path.exists(file_path):
            os.replace(temp_file, file_path)
        else:
            os.rename(temp_file, file_path)
        return True
    except Exception as e:
        print(f"Write failed: {e}")
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except:
            pass
        return False

# ДОПОМІЖНІ ФУНКЦІЇ
def get_current_time_kiev():
    """Отримання поточного часу у Києві"""
    kiev_tz = pytz.timezone('Europe/Kiev')
    now = datetime.now(kiev_tz)
    return now.strftime("%H:%M; %d/%m/%Y")

def load_muted_users_from_file():
    """Завантаження списку замучених користувачів"""
    data = safe_json_read(DATA_FILE)
    muted_users = {}

    for user in data.get("users", []):
        try:
            if user.get("mute", False):
                mute_end = user.get("mute_end")
                if mute_end:
                    try:
                        mute_end = datetime.strptime(mute_end, "%H:%M; %d/%m/%Y")
                    except:
                        mute_end = None

                muted_users[user["id"]] = {
                    "first_name": user.get("first_name", ""),
                    "username": user.get("username", ""),
                    "expiration": mute_end,
                    "reason": user.get("reason", "")
                }
        except KeyError as e:
            print(f"Invalid user format: {e}")
            continue

    return muted_users

def load_sent_messages():
    """Завантаження відправлених повідомлень"""
    data = safe_json_read(DATA_FILE)
    return data.get("sent_messages", {})

def save_sent_messages(sent_messages):
    """Збереження відправлених повідомлень"""
    data = safe_json_read(DATA_FILE)
    data["sent_messages"] = sent_messages
    safe_json_write(data, DATA_FILE)

def load_users_info():
    """Завантаження інформації про користувачів"""
    data = safe_json_read(DATA_FILE)
    return data.get("users", [])

def load_chat_id_from_file():
    """Завантаження ID чату"""
    data = safe_json_read(DATA_FILE)
    return data.get("chat_id", "")

def load_bottocen_from_file():
    """Завантаження токену бота"""
    data = safe_json_read(DATA_FILE)
    return data.get("bot_token", "")


def load_allusers_tem_id_from_file():
    """Завантаження ID теми для всіх користувачів"""
    data = safe_json_read(DATA_FILE)
    return data.get("allusers_tem_id", 386)  # Значення за замовчуванням 386

def load_cave_chat_id_from_file():
    """Завантаження ID печерного чату"""
    data = safe_json_read(DATA_FILE)
    return data.get("cave_chat_id", -1002648725095)  # Значення за замовчуванням -1002648725095




def is_programmer(username):
    """Перевірка, чи є користувач програмістом"""
    data = safe_json_read(DATA_FILE)
    return username in data.get("programmers", [])

def is_admin(username):
    """Перевірка, чи є користувач адміністратором"""
    data = safe_json_read(DATA_FILE)
    return username in data.get("admins", [])


# КОНСТАНТИ ТА НАЛАШТУВАННЯ
DATA_FILE = "data.json"
application = None
app = Flask(__name__)
CREATOR_CHAT_ID = load_chat_id_from_file()  # ID чату для адміністраторів
ALLUSERS_TEM_ID=load_allusers_tem_id_from_file()
CAVE_CHAT_ID= load_cave_chat_id_from_file()


BOTTOCEN = load_bottocen_from_file()

@app.route("/")
def index():
    return "@Supp0rtsBot"

def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)

# ФУНКЦІЇ ДЛЯ РОБОТИ З EXCEL
async def export_to_excel():
    """Експорт даних у Excel файл"""
    try:
        data = safe_json_read(DATA_FILE)
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        excel_filename = f"SupportBot_{current_time}.xlsx"

        all_users_df = pd.DataFrame(data["users"])
        banned_ids = set(data["banned_users"].keys())

        all_users_df = all_users_df.rename(columns={
            'mute': 'mute/ban',
            'mute_end': 'mute/ban_end'
        })

        for user_id in banned_ids:
            mask = all_users_df['id'] == user_id
            all_users_df.loc[mask, 'mute/ban'] = True
            all_users_df.loc[mask, 'mute/ban_end'] = "Навсегда (бан)"

        users_df = all_users_df[all_users_df["mute/ban"] == False].copy()
        muted_df = all_users_df[(all_users_df["mute/ban"] == True) & (~all_users_df['id'].isin(banned_ids))].copy()

        banned_users_info = []
        for user_id, ban_info in data["banned_users"].items():
            user_data = all_users_df[all_users_df["id"] == user_id].iloc[0].to_dict() if not all_users_df[
                all_users_df["id"] == user_id].empty else {
                "username": "Unknown",
                "first_name": "Unknown",
                "join_date": "",
                "rating": 0,
                "mute/ban": True,
                "mute/ban_end": "Навсегда (бан)"
            }

            banned_users_info.append({
                "id": user_id,
                "username": user_data.get("username", "Unknown"),
                "first_name": user_data.get("first_name", "Unknown"),
                "join_date": user_data.get("join_date", ""),
                "rating": user_data.get("rating", 0),
                "mute/ban": True,
                "mute/ban_end": "Навсегда (бан)",
                "reason": ban_info.get("reason", "Banned")
            })

        banned_df = pd.DataFrame(banned_users_info)
        topics_df = pd.DataFrame({
            "user_id": list(data.get("topics", {}).keys()),
            "topic_id": list(data.get("topics", {}).values())
        })

        user_topics_df = pd.DataFrame({
            "topic_id": list(data.get("user_topics", {}).keys()),
            "user_id": list(data.get("user_topics", {}).values())
        })

        sent_messages_df = pd.DataFrame([
            {"message_id": k, "user_id": v}
            for k, v in data.get("sent_messages", {}).items()
        ])

        def process_dates(df):
            if "mute/ban_end" in df.columns:
                df.loc[:, "mute/ban_end"] = df["mute/ban_end"].apply(
                    lambda x: x if "Навсегда" in str(x) else
                    (datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y")
                     if isinstance(x, str) and x != "Навсегда" else "")
                )
            if "join_date" in df.columns:
                df.loc[:, "join_date"] = df["join_date"].apply(
                    lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y")
                    if isinstance(x, str) else ""
                )
            return df

        all_users_df = process_dates(all_users_df)
        users_df = process_dates(users_df)
        muted_df = process_dates(muted_df)
        banned_df = process_dates(banned_df)

        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            all_users_df.to_excel(writer, index=False, sheet_name="AllUsers")
            users_df.to_excel(writer, index=False, sheet_name="ActiveUsers")
            muted_df.to_excel(writer, index=False, sheet_name="MutedUsers")
            banned_df.to_excel(writer, index=False, sheet_name="BannedUsers")

            topics_df.to_excel(writer, index=False, sheet_name="Topics")
            user_topics_df.to_excel(writer, index=False, sheet_name="UserTopics")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")

            pd.DataFrame(data.get("admins", []), columns=["Admins"]).to_excel(
                writer, index=False, sheet_name="Admins")
            pd.DataFrame(data.get("programmers", []), columns=["Programmers"]).to_excel(
                writer, index=False, sheet_name="Programmers")

            # Обновленный GeneralInfo с новыми полями в нужном порядке
            pd.DataFrame([{
                "bot_token": data.get("bot_token", ""),
                "owner_id": data.get("owner_id", ""),
                "chat_id": data.get("chat_id", ""),
                "cave_chat_id": data.get("cave_chat_id", "-1002648725095"),  # Строка по умолчанию
                "allusers_tem_id": data.get("allusers_tem_id", 386),  # Число по умолчанию
                "total_score": data.get("total_score", 0),
                "num_of_ratings": data.get("num_of_ratings", 0)
            }]).to_excel(writer, index=False, sheet_name="GeneralInfo")

            workbook = writer.book
            light_blue_fill = PatternFill(start_color="8bbef2", end_color="8bbef2", fill_type="solid")
            light_green_fill = PatternFill(start_color="8bf28b", end_color="8bf28b", fill_type="solid")
            light_red_fill = PatternFill(start_color="f28b8b", end_color="f28b8b", fill_type="solid")
            light_yellow_fill = PatternFill(start_color="f2f28b", end_color="f2f28b", fill_type="solid")

            admins = data.get("admins", [])
            programmers = data.get("programmers", [])

            if "AllUsers" in workbook.sheetnames:
                ws = workbook["AllUsers"]
                for row in ws.iter_rows(min_row=2):
                    user_id = row[0].value
                    username = row[1].value if len(row) > 1 else ""

                    if str(user_id) in banned_ids:
                        for cell in row:
                            cell.fill = light_red_fill
                    elif ws.cell(row=row[0].row, column=6).value == True:
                        for cell in row:
                            cell.fill = light_yellow_fill
                    elif username in programmers:
                        for cell in row:
                            cell.fill = light_green_fill
                    elif username in admins:
                        for cell in row:
                            cell.fill = light_blue_fill

            if "ActiveUsers" in workbook.sheetnames:
                ws = workbook["ActiveUsers"]
                for row in ws.iter_rows(min_row=2):
                    username = row[1].value if len(row) > 1 else ""

                    if username in programmers:
                        for cell in row:
                            cell.fill = light_green_fill
                    elif username in admins:
                        for cell in row:
                            cell.fill = light_blue_fill

            if "MutedUsers" in workbook.sheetnames:
                ws = workbook["MutedUsers"]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = light_yellow_fill

            if "BannedUsers" in workbook.sheetnames:
                ws = workbook["BannedUsers"]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = light_red_fill

        return excel_filename

    except Exception as e:
        logging.error(f"Помилка при експорті в Excel: {e}")
        return None

async def import_from_excel(file_path):
    """Імпорт даних з Excel файлу"""
    try:
        data = safe_json_read(DATA_FILE)
        new_data = {
            "users": [],
            "muted_users": {},
            "banned_users": {},
            "admins": data.get("admins", []),
            "programmers": data.get("programmers", []),
            "bot_token": data.get("bot_token", ""),
            "owner_id": data.get("owner_id", ""),
            "chat_id": data.get("chat_id", ""),
            "cave_chat_id": data.get("cave_chat_id", "-1002648725095"),  # Строка по умолчанию
            "allusers_tem_id": data.get("allusers_tem_id", 386),  # Число по умолчанию
            "total_score": data.get("total_score", 0),
            "num_of_ratings": data.get("num_of_ratings", 0),
            "sent_messages": {},
            "topics": {},
            "user_topics": {}
        }

        wb = load_workbook(file_path)

        # Обновленный блок для чтения GeneralInfo с новыми полями в нужном порядке
        if "GeneralInfo" in wb.sheetnames:
            ws = wb["GeneralInfo"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 7:
                    if len(headers) >= 1 and row[0]:
                        new_data["bot_token"] = str(row[0])
                    if len(headers) >= 2 and row[1]:
                        new_data["owner_id"] = str(row[1])
                    if len(headers) >= 3 and row[2]:
                        new_data["chat_id"] = str(row[2])
                    if len(headers) >= 4 and row[3]:
                        new_data["cave_chat_id"] = str(row[3])  # Сохраняем как строку
                    if len(headers) >= 5 and row[4] is not None:
                        new_data["allusers_tem_id"] = int(row[4])  # Сохраняем как число
                    if len(headers) >= 6 and row[5] is not None:
                        new_data["total_score"] = float(row[5])
                    if len(headers) >= 7 and row[6] is not None:
                        new_data["num_of_ratings"] = int(row[6])

        if "BannedUsers" in wb.sheetnames:
            ws = wb["BannedUsers"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3 and len(headers) >= 3:
                    user_id = str(row[0])
                    reason = row[headers.index("reason")] if "reason" in headers else "Імпортовано з файлу"

                    new_data["banned_users"][user_id] = {
                        "reason": reason,
                        "date": get_current_time_kiev()
                    }

        if "AllUsers" in wb.sheetnames:
            ws = wb["AllUsers"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7 and len(headers) >= 7:
                    user_data = dict(zip(headers[:7], row[:7]))

                    if "mute/ban" in user_data:
                        user_data["mute"] = user_data.pop("mute/ban")

                    if "mute/ban_end" in user_data:
                        user_data["mute_end"] = user_data.pop("mute/ban_end")
                        if "Навсегда (бан)" in str(user_data["mute_end"]):
                            user_data["mute_end"] = "Навсегда"

                    if user_data.get("mute", False) and user_data["id"] not in new_data["banned_users"]:
                        reason = user_data.get("reason", "Причина не вказана")
                        new_data["muted_users"][user_data["id"]] = {
                            "expiration": user_data.get("mute_end"),
                            "reason": reason if reason else "Причина не вказана"
                        }

                    new_data["users"].append(user_data)

        if "Topics" in wb.sheetnames:
            ws = wb["Topics"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["topics"][str(row[0])] = row[1]

        if "UserTopics" in wb.sheetnames:
            ws = wb["UserTopics"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["user_topics"][str(row[0])] = str(row[1])

        if "SentMessages" in wb.sheetnames:
            ws = wb["SentMessages"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["sent_messages"][str(row[0])] = str(row[1])

        if "Admins" in wb.sheetnames:
            ws = wb["Admins"]
            new_data["admins"] = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]

        if "Programmers" in wb.sheetnames:
            ws = wb["Programmers"]
            new_data["programmers"] = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]

        safe_json_write(new_data, DATA_FILE)
        return True

    except Exception as e:
        print(f"Помилка при імпорті з Excel: {e}")
        return False

async def auto_delete_message(bot, chat_id, message_id, delay):
    """Автоматичне видалення повідомлення через заданий час"""
    try:
        await asyncio.sleep(delay)
        await bot.delete_message(chat_id=chat_id, message_id=message_id)
    except telegram.error.BadRequest as e:
        if "message to delete not found" not in str(e):
            print(f"Error deleting message: {e}")
    except Exception as e:
        print(f"Unexpected error deleting message: {e}")

# ОСНОВНІ КОМАНДИ БОТА
async def start(update: Update, context):
    """Обробка команди /start - запуск бота"""
    try:
        user = update.message.from_user
        chat_id = update.effective_chat.id

        if chat_id == CREATOR_CHAT_ID:
            await update.message.reply_text("Команда /start недоступна в цій групі.")
            return

        config = safe_json_read(DATA_FILE)
        user_found = False

        for u in config["users"]:
            if u["id"] == str(user.id):
                user_found = True
                break

        if not user_found:
            new_user = {
                "id": str(user.id),
                "username": user.username or "Не вказано",
                "first_name": user.first_name or "Не вказано",
                "join_date": get_current_time_kiev(),
                "rating": 0,
                "mute": False,
                "mute_end": None,
                "reason": None
            }
            config["users"].append(new_user)
            safe_json_write(config, DATA_FILE)

            # Создаём тему для нового пользователя
            topic_id = await get_or_create_topic(context, user.id, user.first_name)
            if topic_id:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text=f"Дякую за те що завітали в Supp0rtsBot!"
                )

        keyboard = [
            ["/start", "/rate"],
            ["/message", "/stopmessage"],
            ["/fromus", "/help"],
        ]

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

        await update.message.reply_text(
            "👋 Привіт! Я ваш бот підтримки. \n"
            "📝 Введіть команду /rate для оцінки бота, \n"
            "✉️ /message для написання адміністраторам \n"
            "❓ або /help для допомоги.",
            reply_markup=reply_markup
        )
    except Exception as e:
        print(f"Помилка в start: {e}")
        await update.message.reply_text("Сталася помилка. Спробуйте ще раз.")

async def rate(update: Update, context):
    """Обробка команди /rate - оцінка бота"""
    try:
        user_id = update.message.from_user.id
        data = safe_json_read(DATA_FILE)

        user_rating = None
        for user in data.get("users", []):
            if user.get('id') == str(user_id):
                user_rating = user['rating']
                break

        total_score = data.get("total_score", 0)
        num_of_ratings = data.get("num_of_ratings", 0)
        average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

        rating_text = f"🌟 Загальна оцінка: {round(average_rating, 1)}⭐️\nВаш попередній відгук: {user_rating}⭐️" if user_rating else f"🌟 Загальна оцінка: {round(average_rating, 1)}⭐️"

        keyboard = [
            [InlineKeyboardButton("0.5⭐️", callback_data='0.5'), InlineKeyboardButton("1⭐️", callback_data='1')],
            [InlineKeyboardButton("1.5⭐️", callback_data='1.5'), InlineKeyboardButton("2⭐️", callback_data='2')],
            [InlineKeyboardButton("2.5⭐️", callback_data='2.5'), InlineKeyboardButton("3⭐️", callback_data='3')],
            [InlineKeyboardButton("3.5⭐️", callback_data='3.5'), InlineKeyboardButton("4⭐️", callback_data='4')],
            [InlineKeyboardButton("4.5⭐️", callback_data='4.5'), InlineKeyboardButton("5⭐️", callback_data='5')],
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(f"{rating_text}\nОберіть оцінку:", reply_markup=reply_markup)
    except Exception as e:
        print(f"Помилка в rate: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def button_callback(update: Update, context):
    """Обробка натискання кнопок при оцінюванні"""
    try:
        query = update.callback_query
        await query.answer()
        user_id = query.from_user.id
        new_rating = float(query.data)

        data = safe_json_read(DATA_FILE)
        user_found = False
        previous_rating = 0

        for user in data.get("users", []):
            if user.get('id') == str(user_id):
                previous_rating = user.get('rating', 0)
                user['rating'] = new_rating
                user_found = True
                break

        if not user_found:
            new_user = {
                'id': str(user_id),
                'first_name': query.from_user.first_name,
                'username': query.from_user.username,
                'join_date': get_current_time_kiev(),
                'rating': new_rating,
                'mute': False,
                'mute_end': None,
                'reason': None
            }
            data['users'].append(new_user)

        total_score = data.get("total_score", 0)
        num_of_ratings = data.get("num_of_ratings", 0)

        if previous_rating == 0:
            num_of_ratings += 1
            total_score += new_rating
        else:
            total_score = total_score - previous_rating + new_rating

        data["total_score"] = total_score
        data["num_of_ratings"] = num_of_ratings

        safe_json_write(data, DATA_FILE)

        average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0
        await query.edit_message_text(
            f"🙏 Дякуємо за ваш відгук! Ваша оцінка: {new_rating}⭐️\n"
            f"🌟 Загальна оцінка: {round(average_rating, 1)}⭐️"
        )
    except Exception as e:
        print(f"Помилка в button_callback: {e}")
        await query.edit_message_text("Сталася помилка при обробці вашого відгуку.")

async def message(update: Update, context):
    """Обробка команди /message - надсилання повідомлення адмінам"""
    try:
        muted_users = load_muted_users_from_file()
        user_id = update.message.from_user.id

        if str(user_id) in muted_users:
            mute_info = muted_users[str(user_id)]
            if mute_info['expiration'] and mute_info['expiration'] > datetime.now():
                reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        context.user_data['waiting_for_message'] = True
        reply = await update.message.reply_text(
            "📩 Введіть ваше повідомлення, і його буде відправлено адміністраторам бота. \n"
            "🚫 Введіть /stopmessage, щоб завершити введення повідомлень."
        )
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    except Exception as e:
        print(f"Помилка в команді message: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def stopmessage(update: Update, context):
    """Обробка команди /stopmessage - завершення введення повідомлень"""
    try:
        if context.user_data.get('waiting_for_message'):
            reply = await update.message.reply_text("✅ Ви завершили введення повідомлень.")
            context.user_data['waiting_for_message'] = False
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
        else:
            await update.message.reply_text("Ви не в режимі введення повідомлень.")
    except Exception as e:
        print(f"Помилка в stopmessage: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def help(update: Update, context):
    """Обробка команди /help - довідка про команди"""
    try:
        if str(update.message.chat.id) == str(CREATOR_CHAT_ID):
            help_text = (
                "Доступні команди в групі:\n"
                "Відповісти на повідомлення бота - Надіслати повідомлення користувачу, який надіслав це повідомлення.\n"
                "/mute <час> <користувач> 'причина' - Замутити користувача на вказаний час.\n"
                "/unmute <користувач> - Розмутити користувача.\n"
                "/mutelist - Показати список замучених користувачів.\n"
                "/alllist - Показати всіх користувачів.\n"
                "/fromus - Інформація про створювача.\n"
                "/help - Показати доступні команди.\n"
                "/info - Показати інформацію про програмістів та адміністраторів.\n"
                "/admin <користувач> - Додати адміністратора.\n"
                "/deleteadmin <користувач> - Видалити адміністратора.\n"
                "/programier <користувач> - Додати програміста.\n"
                "/deleteprogramier <користувач> - Видалити програміста.\n"
                "/get_alllist - Отримати Excel файл з користувачами.\n"
                "/set_alllist - Записати Excel файл з користувачами.\n"
            )
        else:
            help_text = (
                "ℹ️ Доступні команди в боті:\n"
                "🚀 /start - Запустити бота\n"
                "⭐️ /rate - Залишити відгук\n"
                "✉️ /message - Почати введення повідомлень адміністраторам\n"
                "🚫 /stopmessage - Завершити введення повідомлень\n"
                "👨‍💻 /fromus - Інформація про створювача\n"
                "❓ /help - Показати доступні команди"
            )

        await update.message.reply_text(help_text)
    except Exception as e:
        print(f"Помилка в help: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def fromus(update: Update, context):
    """Обробка команди /fromus - інформація про творця"""
    try:
        await update.message.reply_text(
            "👨‍💻 *Skeleton* Написав бота\n"
            "📂 Портфоліо: ```https://www.linkedin.com/in/artem-k-972a41344/```\n"
            "📢 Телеграм канал з усіма проєктами: ```https://t.me/AboutMyProjects```\n"
            "❓ По всім питанням пишіть в цього бота",
            parse_mode="MarkdownV2"
        )
    except Exception as e:
        print(f"Помилка в fromus: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def info(update: Update, context: CallbackContext):
    """Обробка команди /info - інформація про адміністраторів"""
    try:
        data = safe_json_read(DATA_FILE)
        programmers = data.get("programmers", [])
        admins = data.get("admins", [])

        programmer_list = "\n".join(programmers) if programmers else "Список програмістів пустий."
        admin_list = "\n".join(admins) if admins else "Список адміністраторів пустий."

        await update.message.reply_text(f"👨‍💻 Програмісти:\n{programmer_list}\n\n👮 Адміністратори:\n{admin_list}")
    except Exception as e:
        print(f"Помилка в info: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

# КОМАНДИ АДМІНІСТРАТОРІВ
async def mute(update: Update, context: CallbackContext):
    """Обробка команди /mute - обмеження користувача"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна лише адміністраторам.")
            return

        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Ця команда працює лише в темах користувачів.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не вдалося визначити користувача для цієї теми.")
            return

        if user_id in data["banned_users"]:
            await update.message.reply_text("❌ Цей користувач забанений і не може бути замучений!")
            return

        mute_time = 300
        reason = "За рішенням адміністрації"

        if context.args:
            if context.args[0].isdigit():
                mute_time = int(context.args[0])
                if len(context.args) > 1:
                    reason = ' '.join(context.args[1:])
            else:
                reason = ' '.join(context.args)

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Користувача не знайдено.")
            return

        if user_data["id"] == data["owner_id"]:
            await update.message.reply_text("Неможливо замутити власника чату.")
            return

        mute_end = (datetime.now() + timedelta(seconds=mute_time)).strftime("%H:%M; %d/%m/%Y")
        user_data.update({
            "mute": True,
            "mute_end": mute_end,
            "reason": reason
        })

        data["muted_users"][user_id] = {
            "expiration": mute_end,
            "reason": reason
        }

        safe_json_write(data, DATA_FILE)

        mute_permissions = ChatPermissions(
            can_send_messages=False,
            can_send_photos=False,
            can_send_videos=False,
            can_send_audios=False,
            can_send_documents=False,
            can_send_polls=False,
            can_send_other_messages=False,
            can_add_web_page_previews=False,
            can_change_info=False,
            can_invite_users=False,
            can_pin_messages=False
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=mute_permissions,
            until_date=int((datetime.now() + timedelta(seconds=mute_time)).timestamp())
        )

        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"🔇 Вас замутили на {mute_time} секунд\n"
                     f"📌 Причина: {reason}\n"
                     f"⏳ Мут закінчиться: {mute_end}"
            )
        except Exception as e:
            print(f"Помилка сповіщення про мут: {e}")

        await update.message.reply_text(f"✅ Користувача замучено на {mute_time} секунд. Причина: {reason}")

    except Exception as e:
        print(f"Помилка в команді mute: {e}")
        await update.message.reply_text("❌ Сталася помилка при обробці команди.")

async def unmute(update: Update, context: CallbackContext):
    """Обробка команди /unmute - зняття обмежень"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна лише адміністраторам.")
            return

        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Ця команда працює лише в темах користувачів.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не вдалося визначити користувача для цієї теми.")
            return

        if user_id in data["banned_users"]:
            await update.message.reply_text("❌ Цей користувач забанений! Використовуйте /unban для розбану.")
            return

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Користувача не знайдено.")
            return

        if not user_data["mute"]:
            await update.message.reply_text("Цей користувач не в муті.")
            return

        user_data.update({
            "mute": False,
            "mute_end": None,
            "reason": None
        })

        if user_id in data["muted_users"]:
            del data["muted_users"][user_id]

        safe_json_write(data, DATA_FILE)

        unmute_permissions = ChatPermissions(
            can_send_messages=True,
            can_send_photos=True,
            can_send_videos=True,
            can_send_audios=True,
            can_send_documents=True,
            can_send_polls=True,
            can_send_other_messages=True,
            can_add_web_page_previews=True,
            can_change_info=True,
            can_invite_users=True,
            can_pin_messages=True
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=unmute_permissions
        )

        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text="🔊 Вас розмутили. Тепер ви знову можете писати в чат."
            )
        except Exception as e:
            print(f"Помилка сповіщення про розмут: {e}")

        await update.message.reply_text(f"✅ Користувача @{user_data['username']} було розмучено.")

    except Exception as e:
        print(f"Помилка в команді unmute: {e}")
        await update.message.reply_text("❌ Сталася помилка при обробці команди.")

async def ban(update: Update, context: CallbackContext):
    """Обробка команди /ban - бан користувача"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна лише адміністраторам.")
            return

        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Ця команда працює лише в темах користувачів.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не вдалося визначити користувача для цієї теми.")
            return

        reason = "За рішенням адміністрації"
        if context.args:
            reason = ' '.join(context.args)

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Користувача не знайдено.")
            return

        if user_data["id"] == data["owner_id"]:
            await update.message.reply_text("Неможливо забанити власника чату.")
            return

        data["banned_users"][user_id] = {
            "reason": reason,
            "date": get_current_time_kiev()
        }

        user_data.update({
            "mute": True,
            "mute_end": "Назавжди",
            "reason": f"Забанен: {reason}"
        })

        data["muted_users"][user_id] = {
            "expiration": "Назавжди",
            "reason": f"Забанен: {reason}"
        }

        safe_json_write(data, DATA_FILE)

        await context.bot.ban_chat_member(
            chat_id=data["chat_id"],
            user_id=int(user_id)
        )

        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"🚫 Вас забанено назавжди\n"
                     f"📌 Причина: {reason}"
            )
        except Exception as e:
            print(f"Помилка сповіщення користувача про бан: {e}")

        await update.message.reply_text(f"Користувача забанено назавжди. Причина: {reason}")

    except Exception as e:
        print(f"Помилка в команді ban: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def unban(update: Update, context: CallbackContext):
    """Обробка команди /unban - розбан користувача"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна лише адміністраторам.")
            return

        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Ця команда працює лише в темах користувачів.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не вдалося визначити користувача для цієї теми.")
            return

        if user_id not in data["banned_users"]:
            await update.message.reply_text("Цей користувач не забанений.")
            return

        del data["banned_users"][user_id]

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if user_data:
            user_data.update({
                "mute": False,
                "mute_end": None,
                "reason": None
            })

        if user_id in data["muted_users"]:
            del data["muted_users"][user_id]

        safe_json_write(data, DATA_FILE)

        await context.bot.unban_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id)
        )

        unmute_permissions = ChatPermissions(
            can_send_messages=True,
            can_send_photos=True,
            can_send_videos=True,
            can_send_audios=True,
            can_send_documents=True,
            can_send_polls=True,
            can_send_other_messages=True,
            can_add_web_page_previews=True,
            can_change_info=True,
            can_invite_users=True,
            can_pin_messages=True
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=unmute_permissions
        )

        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text="✅ Вас розбанено. Тепер ви знову можете брати участь у чаті."
            )
        except Exception as e:
            print(f"Помилка сповіщення про розбан: {e}")

        await update.message.reply_text(f"✅ Користувача розбанено.")

    except Exception as e:
        print(f"Помилка в команді unban: {e}")
        await update.message.reply_text("❌ Сталася помилка при обробці команди.")

async def check_mute_expirations():
    """Перевірка закінчення часу муту"""
    try:
        application = Application.builder().token(BOTTOCEN).build()
        async with application:
            context = ContextTypes.DEFAULT_TYPE(application=application)

            data = safe_json_read(DATA_FILE)
            now = datetime.now()
            users_to_unmute = []

            for user in data["users"]:
                if user["mute"] and user["mute_end"]:
                    try:
                        mute_end = datetime.strptime(user["mute_end"], "%H:%M; %d/%m/%Y")
                        if mute_end <= now:
                            users_to_unmute.append(user)
                    except ValueError:
                        continue

            if users_to_unmute:
                for user in users_to_unmute:
                    user.update({
                        "mute": False,
                        "mute_end": None,
                        "reason": None
                    })

                    if "muted_users" in data and user["id"] in data["muted_users"]:
                        del data["muted_users"][user["id"]]

                    try:
                        await context.bot.restrict_chat_member(
                            chat_id=int(data["chat_id"]),
                            user_id=int(user["id"]),
                            permissions=ChatPermissions(
                                can_send_messages=True,
                                can_send_photos=True,
                                can_send_videos=True,
                                can_send_audios=True,
                                can_send_documents=True,
                                can_send_polls=True,
                                can_send_other_messages=True,
                                can_add_web_page_previews=True,
                                can_change_info=True,
                                can_invite_users=True,
                                can_pin_messages=True
                            )
                        )

                        try:
                            await context.bot.send_message(
                                chat_id=int(user["id"]),
                                text="🔊 Ваш мут закінчився. Тепер ви знову можете писати в чат."
                            )
                        except Exception as e:
                            print(f"Помилка сповіщення користувача про закінчення муту: {e}")

                    except Exception as e:
                        print(f"Помилка при розмуті користувача {user['id']}: {e}")

                print(f"Дані перед збереженням: {data['users']}")
                if not safe_json_write(data, DATA_FILE):
                    print("Помилка збереження даних")
                else:
                    print("Дані успішно збережено")
                    check_data = safe_json_read(DATA_FILE)
                    print(f"Дані після збереження: {check_data['users']}")

    except Exception as e:
        print(f"Помилка в перевірці строків муту: {e}")

async def admin(update: Update, context: CallbackContext):
    """Обробка команди /admin - додавання адміністратора"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /admin @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["admins"]:
            await update.message.reply_text(f"Користувач @{username} вже є адміністратором.")
        else:
            data["admins"].append(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"👮 Користувач @{username} доданий до списку адміністраторів.")
    except Exception as e:
        print(f"Помилка в admin: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def deleteadmin(update: Update, context: CallbackContext):
    """Обробка команди /deleteadmin - видалення адміністратора"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /deleteadmin @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["admins"]:
            data["admins"].remove(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"👮 Користувач @{username} видалений зі списку адміністраторів.")
        else:
            await update.message.reply_text(f"Користувач @{username} не знайдений.")
    except Exception as e:
        print(f"Помилка в deleteadmin: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def programier(update: Update, context: CallbackContext):
    """Обробка команди /programier - додавання програміста"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /programier @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["programmers"]:
            await update.message.reply_text(f"Користувач @{username} вже є програмістом.")
        else:
            data["programmers"].append(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"👨‍💻 Користувач @{username} доданий до списку програмістів.")
    except Exception as e:
        print(f"Помилка в programier: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def deleteprogramier(update: Update, context: CallbackContext):
    """Обробка команди /deleteprogramier - видалення програміста"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /deleteprogramier @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username == "ArtemKirss":
            await update.message.reply_text(f"Неможливо видалити {username} зі списку програмістів.")
        elif username in data["programmers"]:
            data["programmers"].remove(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"👨‍💻 Користувач @{username} видалений зі списку програмістів.")
        else:
            await update.message.reply_text(f"Користувач @{username} не є програмістом.")
    except Exception as e:
        print(f"Помилка в deleteprogramier: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def mutelist(update: Update, context):
    """Обробка команди /mutelist - список замучених користувачів"""
    try:
        user = update.message.from_user.username
        if update.message.chat.id != CREATOR_CHAT_ID:
            if not is_programmer(user) and not is_admin(user):
                reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        data = safe_json_read(DATA_FILE)
        admins = data.get("admins", [])
        programmers = data.get("programmers", [])
        muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

        response = "Замучені користувачі:\n"

        if muted_users:
            for user_id, mute_info in muted_users.items():
                expiration = mute_info.get('mute_end', 'Невідомо')
                reason = mute_info.get('reason', 'Без причини')

                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"

                join_date = mute_info.get('join_date', 'Невідома')
                rating = mute_info.get('rating', 0)
                mute_symbol = "🔇"

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                response += (
                    f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                    f"Залишилось: {expiration}\n"
                    f"Причина: {reason}\n"
                    f"Дата заходу: {join_date}\n"
                    f"Оцінка: {rating}⭐️\n"
                    "-------------------------------------------------------------------------\n"
                )
        else:
            response += "Немає замучених користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        await update.message.reply_text(response)
    except Exception as e:
        print(f"Помилка в mutelist: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def alllist(update: Update, context: CallbackContext):
    """Обробка команди /alllist - список всіх користувачів"""
    try:
        user = update.message.from_user.username
        if update.message.chat.id != CREATOR_CHAT_ID:
            if not is_programmer(user) and not is_admin(user):
                reply = await update.message.reply_text("Ця команда доступна лише адміністраторам бота.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        data = safe_json_read(DATA_FILE)
        admins = data.get("admins", [])
        programmers = data.get("programmers", [])
        users_info = {user['id']: user for user in data.get("users", [])}
        muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

        response = "Користувачі:\n"
        unique_users = {user['id'] for user in data.get("users", [])}

        if unique_users:
            for user_id in unique_users:
                user_data = users_info.get(str(user_id), {})
                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"
                join_date = user_data.get('join_date', 'Невідома')
                rating = user_data.get('rating', 0)

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                mute_symbol = "🔇" if str(user_id) in muted_users else "🔊"

                response += f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\nДата заходу: {join_date}\nОцінка: {rating}⭐️\n"
                response += "-------------------------------------------------------------------------\n"
        else:
            response += "Немає користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        response += "==========================================\n"
        response += "\n"
        response += "==========================================\n"
        response += "Замучені користувачі:\n"

        if muted_users:
            for user_id, mute_info in muted_users.items():
                expiration = mute_info['mute_end'] or "Невідомо"
                reason = mute_info.get('reason', "Без причини")
                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"
                user_data = users_info.get(str(user_id), {})
                join_date = user_data.get('join_date', 'Невідома')
                rating = user_data.get('rating', 0)

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                mute_symbol = "🔇"

                response += (
                    f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                    f"Залишилось: {str(expiration).split('.')[0]}\n"
                    f"Причина: {reason}\n"
                    f"Дата заходу: {join_date}\n"
                    f"Оцінка: {rating}⭐️\n"
                    "-------------------------------------------------------------------------\n"
                )
        else:
            response += "Немає замучених користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        await update.message.reply_text(response)
    except Exception as e:
        print(f"Помилка в alllist: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def get_alllist(update: Update, context: CallbackContext) -> None:
    """Обробка команди /get_alllist - отримання Excel файлу з даними"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
            return

        excel_filename = await export_to_excel()
        if excel_filename:
            with open(excel_filename, "rb") as file:
                filename_to_send = os.path.basename(excel_filename)
                await update.message.reply_document(
                    document=file,
                    filename=filename_to_send
                )
            try:
                os.remove(excel_filename)
            except Exception as e:
                print(f"Помилка при видаленні файлу: {e}")
        else:
            await update.message.reply_text("Помилка при створенні Excel-файлу")
    except Exception as e:
        print(f"Помилка в get_alllist: {e}")
        await update.message.reply_text("Сталася помилка при експорті даних")

async def set_alllist(update: Update, context: CallbackContext) -> None:
    """Обробка команди /set_alllist - імпорт даних з Excel файлу"""
    try:
        user = update.message.from_user.username

        await update.message.reply_text("Будь ласка, надішліть Excel-файл з даними.")
        context.user_data["awaiting_file"] = True
    except Exception as e:
        print(f"Помилка в set_alllist: {e}")
        await update.message.reply_text("Сталася помилка")

async def get_logs(update: Update, context: CallbackContext):
    """Обробка команди /get_logs - отримання логів"""
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна лише адміністраторам.")
            return

        log_file = "bot_errors.log"

        if not os.path.exists(log_file):
            await update.message.reply_text("Файл логів не знайдено.")
            return

        with open(log_file, "rb") as file:
            await update.message.reply_document(
                document=file,
                filename=f"bot_logs_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
            )

    except Exception as e:
        print(f"Помилка при відправці логів: {e}")
        await update.message.reply_text("❌ Сталася помилка при відправці логів.")

# ОБРОБКА ПОВІДОМЛЕНЬ
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обробка всіх повідомлень"""
    try:
        sent_messages = load_sent_messages()
        muted_users = load_muted_users_from_file()
        data = safe_json_read(DATA_FILE)

        if context.user_data.get("awaiting_file"):
            if update.message.document:
                file = await update.message.document.get_file()
                await file.download_to_drive("temp_import.xlsx")

                if await import_from_excel("temp_import.xlsx"):
                    await update.message.reply_text("Дані успішно імпортовано!")
                else:
                    await update.message.reply_text("Помилка при імпорті даних")

                context.user_data["awaiting_file"] = False
                try:
                    os.remove("temp_import.xlsx")
                except:
                    pass
                return

        if update.message.message_thread_id == ALLUSERS_TEM_ID and is_programmer(update.message.from_user.username):
            user = update.message.from_user.username
            if is_programmer(user) or is_admin(user):
                success_count = 0
                fail_count = 0

                for user_data in data.get("users", []):
                    user_id = user_data.get("id")
                    if user_id:
                        try:
                            if update.message.text:
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text=f"📢 <b>Оголошення від адміністрації:</b>\n{update.message.text}",
                                    parse_mode='HTML'
                                )
                            elif update.message.photo:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_photo(
                                    chat_id=int(user_id),
                                    photo=update.message.photo[-1].file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                            elif update.message.document:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_document(
                                    chat_id=int(user_id),
                                    document=update.message.document.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                            elif update.message.sticker:
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text="📢 <b>Оголошення від адміністрації:</b>",
                                    parse_mode='HTML'
                                )
                                await context.bot.send_sticker(
                                    chat_id=int(user_id),
                                    sticker=update.message.sticker.file_id
                                )
                            elif update.message.voice:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_voice(
                                    chat_id=int(user_id),
                                    voice=update.message.voice.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                            elif update.message.video:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_video(
                                    chat_id=int(user_id),
                                    video=update.message.video.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                            elif update.message.video_note:
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text="📢 <b>Оголошення від адміністрації:</b>",
                                    parse_mode='HTML'
                                )
                                await context.bot.send_video_note(
                                    chat_id=int(user_id),
                                    video_note=update.message.video_note.file_id
                                )

                            success_count += 1
                        except Exception as e:
                            print(f"Помилка відправки повідомлення користувачу {user_id}: {str(e)}")
                            fail_count += 1

                report_message = (
                    f"📊 <b>Результат розсилки:</b>\n"
                    f"• ✅ Відправлено: {success_count}\n"
                    f"• ❌ Не вдалося: {fail_count}\n"
                    f"• 👥 Усього користувачів: {len(data.get('users', []))}"
                )

                await update.message.reply_text(
                    report_message,
                    parse_mode='HTML'
                )
                return

        if update.message.chat.id == int(data["chat_id"]) and update.message.message_thread_id is None:
            return

        if str(update.message.chat.id) != str(data["chat_id"]):
            user_id = update.message.from_user.id
            if str(user_id) in muted_users and muted_users[str(user_id)]['expiration'] and muted_users[str(user_id)][
                'expiration'] > datetime.now():
                reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

            if context.user_data.get('waiting_for_message'):
                user_name = update.effective_user.first_name
                user_username = update.effective_user.username if update.effective_user.username else "немає імені користувача"
                current_time = get_current_time_kiev()

                def escape_markdown(text):
                    if not text:
                        return ""
                    escape_chars = r'_*[]()~`>#+-=|{}.!'
                    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', text)

                topic_id = await get_or_create_topic(context, user_id, user_name)

                if topic_id:
                    base_message = f'📩 Повідомлення від **{escape_markdown(user_name)}**; `@{escape_markdown(user_username)}` `{user_id}`\n⏰ {escape_markdown(current_time)}:'
                    if update.message.text:
                        message_text = f'{base_message}\n{escape_markdown(update.message.text)}'
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.photo:
                        photo_file_id = update.message.photo[-1].file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_photo(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            photo=photo_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.document:
                        document_file_id = update.message.document.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_document(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            document=document_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.sticker:
                        sticker_file_id = update.message.sticker.file_id
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=base_message,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        await context.bot.send_sticker(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            sticker=sticker_file_id
                        )
                        save_sent_messages(sent_messages)
                    elif update.message.voice:
                        voice_file_id = update.message.voice.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_voice(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            voice=voice_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.video:
                        video_file_id = update.message.video.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_video(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            video=video_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.video_note:
                        video_note_file_id = update.message.video_note.file_id
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=base_message,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        await context.bot.send_video_note(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            video_note=video_note_file_id
                        )
                        save_sent_messages(sent_messages)

                    reply = await update.message.reply_text("✅ Ваше повідомлення надіслано адміністраторам бота.")
                    await asyncio.create_task(
                        auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
            else:
                await update.message.reply_text("Введіть /message, щоб надсилати повідомлення адміністраторам бота.")
            return

        if update.message.message_thread_id is not None:
            user = update.message.from_user.username
            if not is_programmer(user) and not is_admin(user):
                return

            user_topics = data.get("user_topics", {})
            user_id = user_topics.get(str(update.message.message_thread_id))

            if user_id:
                try:
                    if update.message.text:
                        await context.bot.send_message(
                            chat_id=int(user_id),
                            text=update.message.text
                        )
                    elif update.message.photo:
                        await context.bot.send_photo(
                            chat_id=int(user_id),
                            photo=update.message.photo[-1].file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.document:
                        await context.bot.send_document(
                            chat_id=int(user_id),
                            document=update.message.document.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.sticker:
                        await context.bot.send_sticker(
                            chat_id=int(user_id),
                            sticker=update.message.sticker.file_id
                        )
                    elif update.message.voice:
                        await context.bot.send_voice(
                            chat_id=int(user_id),
                            voice=update.message.voice.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video:
                        await context.bot.send_video(
                            chat_id=int(user_id),
                            video=update.message.video.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video_note:
                        await context.bot.send_video_note(
                            chat_id=int(user_id),
                            video_note=update.message.video_note.file_id
                        )

                    sent_msg = await update.message.reply_text("Повідомлення відправлено користувачу")
                    asyncio.create_task(auto_delete_message(context.bot, chat_id=sent_msg.chat.id, message_id=sent_msg.message_id, delay=5))
                except Exception as e:
                    await update.message.reply_text(f"Помилка при відправці: {str(e)}")
            return

        if update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id:
            original_message_id = str(update.message.reply_to_message.message_id)
            if original_message_id in sent_messages:
                original_user_id = sent_messages[original_message_id]
                reply_text = update.message.text if update.message.text else ""

                user_name = "Користувач"
                for user_data in data['users']:
                    if str(user_data['id']) == str(original_user_id):
                        user_name = user_data['first_name']
                        break

                try:
                    if update.message.photo:
                        await context.bot.send_photo(
                            chat_id=int(original_user_id),
                            photo=update.message.photo[-1].file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.document:
                        await context.bot.send_document(
                            chat_id=int(original_user_id),
                            document=update.message.document.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.sticker:
                        await context.bot.send_sticker(
                            chat_id=int(original_user_id),
                            sticker=update.message.sticker.file_id
                        )
                    elif update.message.voice:
                        await context.bot.send_voice(
                            chat_id=int(original_user_id),
                            voice=update.message.voice.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video:
                        await context.bot.send_video(
                            chat_id=int(original_user_id),
                            video=update.message.video.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video_note:
                        await context.bot.send_video_note(
                            chat_id=int(original_user_id),
                            video_note=update.message.video_note.file_id
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=int(original_user_id),
                            text=reply_text
                        )

                    await update.message.reply_text(f"Користувачу {user_name} було надіслано повідомлення")
                    sent_messages[str(update.message.message_id)] = update.message.from_user.id
                    save_sent_messages(sent_messages)
                except Exception as e:
                    await update.message.reply_text(f"Помилка при відправці: {str(e)}")
    except Exception as e:
        print(f"Помилка в handle_message: {str(e)}")


async def get_or_create_topic(context: ContextTypes.DEFAULT_TYPE, user_id: int, first_name: str):
    """Створення або отримання теми для користувача з обробкою блокувань"""
    try:
        data = safe_json_read(DATA_FILE)
        chat_id = int(data["chat_id"])
        topics = data.get("topics", {})
        user_topics = data.get("user_topics", {})

        # Проверяем, есть ли тема в нашем словаре
        if str(user_id) in topics:
            topic_id = topics[str(user_id)]
            return topic_id

        # Проверяем, не заблокировал ли пользователь бота
        try:
            await context.bot.send_chat_action(
                chat_id=user_id,
                action=telegram.constants.ChatAction.TYPING
            )
        except telegram.error.Forbidden as e:
            if "bot was blocked by the user" in str(e):
                print(f"User {user_id} blocked the bot")
                return None
            raise

        # Створюємо нову тему
        try:
            topic_name = f"{first_name} ({user_id})"
            print(f"Creating new topic with name: {topic_name}")

            forum_topic = await context.bot.create_forum_topic(
                chat_id=chat_id,
                name=topic_name[:128]
            )
            topic_id = forum_topic.message_thread_id

            try:
                user_info = await context.bot.get_chat(user_id)
                username = user_info.username or "немає username"
                full_name = user_info.full_name or first_name
            except Exception as e:
                print(f"Error getting user info: {e}")
                username = "немає username"
                full_name = first_name

            info_message = (
                f"📌 Інформація про користувача:\n"
                f"👤 Ім'я: {full_name}\n"
                f"🔗 Юзернейм: @{username}\n"
                f"🆔 ID: {user_id}\n"
                f"🗂 ID теми: {topic_id}"
            )

            try:
                msg = await context.bot.send_message(
                    chat_id=chat_id,
                    message_thread_id=topic_id,
                    text=info_message
                )
                await context.bot.pin_chat_message(
                    chat_id=chat_id,
                    message_id=msg.message_id
                )
                print(f"Pinned info message in topic {topic_id}")
            except Exception as e:
                print(f"Error pinning message: {e}")

            # Обновляем данные
            topics[str(user_id)] = topic_id
            user_topics[str(topic_id)] = str(user_id)
            data["topics"] = topics
            data["user_topics"] = user_topics

            if not safe_json_write(data, DATA_FILE):
                print("Failed to save data to JSON file")
            else:
                print("Successfully updated topic data")

            return topic_id

        except telegram.error.BadRequest as e:
            print(f"Telegram BadRequest while creating topic: {e}")
            return None
        except Exception as e:
            print(f"Unexpected error while creating topic: {e}")
            return None

    except Exception as e:
        print(f"Critical error in get_or_create_topic: {e}")
        return None

# НАЛАШТУВАННЯ КОМАНД БОТА
async def set_default_commands(application):
    """Встановлення стандартних команд для звичайних користувачів"""
    try:
        commands = [
            BotCommand("start", "Запустити бота"),
            BotCommand("rate", "Залишити відгук"),
            BotCommand("message", "Почати введення повідомлень адміністраторам"),
            BotCommand("stopmessage", "Завершити введення повідомлень"),
            BotCommand("fromus", "Інформація про створювача"),
            BotCommand("help", "Показати доступні команди"),
        ]
        await application.bot.set_my_commands(commands, scope=BotCommandScopeDefault())
    except Exception as e:
        print(f"Помилка в set_default_commands: {e}")

async def set_creator_commands(application):
    """Встановлення команд для адміністраторів"""
    try:
        commands = [
            BotCommand("mutelist", "Показати список замучених користувачів"),
            BotCommand("mute", "Замутити користувача"),
            BotCommand("unmute", "Розмутити користувача"),
            BotCommand("ban", "Забанити користувача"),
            BotCommand("unban", "Розбанити користувача"),
            BotCommand("alllist", "Показати всіх користувачів"),
            BotCommand("fromus", "Інформація про створювача"),
            BotCommand("help", "Показати доступні команди"),
            BotCommand("info", "Показати інформацію про програмістів та адміністраторів"),
            BotCommand("get_alllist", "Отримати Excel файл з користувачами"),
            BotCommand("set_alllist", "Записати Excel файл з користувачами"),
        ]
        await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CREATOR_CHAT_ID))
    except Exception as e:
        print(f"Помилка в set_creator_commands: {e}")

async def set_save_commands(application):
    """Встановлення команд для адміністраторів"""
    commands = [
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
        BotCommand("get_logs", "Отримати логи"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CAVE_CHAT_ID))

async def send_user_list():
    """Автоматична відправка Excel файлу з користувачами"""
    try:
        excel_filename = await export_to_excel()
        if excel_filename:
            bot = Bot(token=BOTTOCEN)
            with open(excel_filename, "rb") as file:
                filename_to_send = os.path.basename(excel_filename)
                await bot.send_document(
                    chat_id=CAVE_CHAT_ID,
                    document=file,
                    filename=filename_to_send
                )
            try:
                os.remove(excel_filename)
            except Exception as e:
                print(f"Помилка при видаленні файлу: {e}")
    except Exception as e:
        print(f"Помилка при відправці списку користувачів: {e}")
        try:
            bot = Bot(token=BOTTOCEN)
            await bot.send_message(chat_id=CAVE_CHAT_ID, text=f"Помилка при створенні звіту: {e}")
        except:
            pass

# ГОЛОВНА ФУНКЦІЯ
async def main():
    """Головна функція для запуску бота"""
    try:
        application = Application.builder().token(BOTTOCEN).build()

        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("rate", rate))
        application.add_handler(CommandHandler("message", message))
        application.add_handler(CommandHandler("stopmessage", stopmessage))
        application.add_handler(CommandHandler("fromus", fromus))
        application.add_handler(CommandHandler("help", help))
        application.add_handler(CommandHandler("mute", mute))
        application.add_handler(CommandHandler("unmute", unmute))
        application.add_handler(CommandHandler("ban", ban))
        application.add_handler(CommandHandler("unban", unban))
        application.add_handler(CommandHandler("mutelist", mutelist))
        application.add_handler(CommandHandler("alllist", alllist))
        application.add_handler(CommandHandler("admin", admin))
        application.add_handler(CommandHandler("deleteadmin", deleteadmin))
        application.add_handler(CommandHandler("programier", programier))
        application.add_handler(CommandHandler("deleteprogramier", deleteprogramier))
        application.add_handler(CommandHandler("info", info))
        application.add_handler(CommandHandler("get_alllist", get_alllist))
        application.add_handler(CommandHandler("set_alllist", set_alllist))
        application.add_handler(CommandHandler("get_logs", get_logs))

        application.add_handler(CallbackQueryHandler(button_callback))
        application.add_handler(MessageHandler(filters.ALL, handle_message))

        await set_default_commands(application)
        await set_creator_commands(application)
        await set_save_commands(application)

        scheduler = AsyncIOScheduler(timezone=pytz.timezone("Europe/Kyiv"))
        scheduler.add_job(send_user_list, "cron", hour=0, minute=0)
        scheduler.add_job(check_mute_expirations, "interval", minutes=1)
        scheduler.start()

        application.run_polling()
    except Exception as e:
        print(f"Помилка в main: {e}")

if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()
    asyncio.run(main())