import asyncio
import re
import os

import nest_asyncio
import pytz
import threading
import json
import pandas as pd
import telegram.error
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ChatPermissions, \
    BotCommand, BotCommandScopeDefault, BotCommandScopeChat, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, \
    ContextTypes
from datetime import datetime, timedelta
from flask import Flask
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from apscheduler.schedulers.background import BackgroundScheduler

nest_asyncio.apply()

import logging
logging.basicConfig(
    filename='bot_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)



# Константы
DATA_FILE = "data.json"
application = None
app = Flask(__name__)


@app.route("/")
def index():
    return "@Supp0rtsBot"


def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)


# Улучшенные функции работы с JSON
def safe_json_read(file_path):
    default_data = {
        "users": [],
        "muted_users": {},
        "banned_users": {},
        "admins": [],
        "programmers": [],
        "bot_token": "",
        "owner_id": "",
        "chat_id": "",
        "total_score": 0.0,
        "num_of_ratings": 0,
        "sent_messages": {},
        "topics": {},
        "user_topics": {}
    }

    if not os.path.exists(file_path):
        safe_json_write(default_data, file_path)
        return default_data

    try:
        for encoding in ['utf-8-sig', 'utf-8', 'cp1251']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    data = json.load(f)
                    for key in default_data.keys():
                        if key not in data:
                            data[key] = default_data[key]
                    return data
            except UnicodeDecodeError:
                continue
            except json.JSONDecodeError as je:
                print(f"Invalid JSON in {file_path}: {je}")
                continue

        print("All encoding attempts failed, creating new file")
        safe_json_write(default_data, file_path)
        return default_data
    except Exception as e:
        print(f"Critical read error: {e}")
        safe_json_write(default_data, file_path)
        return default_data

def safe_json_write(data, file_path):
    temp_file = file_path + '.tmp'
    try:
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        with open(temp_file, 'r', encoding='utf-8') as f:
            json.load(f)

        if os.path.exists(file_path):
            os.replace(temp_file, file_path)
        else:
            os.rename(temp_file, file_path)
        return True
    except Exception as e:
        print(f"Write failed: {e}")
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except:
            pass
        return False

# Обновленные вспомогательные функции
def get_current_time_kiev():
    kiev_tz = pytz.timezone('Europe/Kiev')
    now = datetime.now(kiev_tz)
    return now.strftime("%H:%M; %d/%m/%Y")

def load_muted_users_from_file():
    data = safe_json_read(DATA_FILE)
    muted_users = {}

    for user in data.get("users", []):
        try:
            if user.get("mute", False):
                mute_end = user.get("mute_end")
                if mute_end:
                    try:
                        mute_end = datetime.strptime(mute_end, "%H:%M; %d/%m/%Y")
                    except:
                        mute_end = None

                muted_users[user["id"]] = {
                    "first_name": user.get("first_name", ""),
                    "username": user.get("username", ""),
                    "expiration": mute_end,
                    "reason": user.get("reason", "")
                }
        except KeyError as e:
            print(f"Invalid user format: {e}")
            continue

    return muted_users


def load_sent_messages():
    data = safe_json_read(DATA_FILE)
    return data.get("sent_messages", {})


def save_sent_messages(sent_messages):
    data = safe_json_read(DATA_FILE)
    data["sent_messages"] = sent_messages
    safe_json_write(data, DATA_FILE)


def load_users_info():
    data = safe_json_read(DATA_FILE)
    return data.get("users", [])


def load_chat_id_from_file():
    data = safe_json_read(DATA_FILE)
    return data.get("chat_id", "")


def load_bottocen_from_file():
    data = safe_json_read(DATA_FILE)
    return data.get("bot_token", "")


def is_programmer(username):
    data = safe_json_read(DATA_FILE)
    return username in data.get("programmers", [])


def is_admin(username):
    data = safe_json_read(DATA_FILE)
    return username in data.get("admins", [])


# Инициализация глобальных переменных
config = safe_json_read(DATA_FILE)
users_info = load_users_info()
muted_users = load_muted_users_from_file()
CREATOR_CHAT_ID = load_chat_id_from_file()
BOTTOCEN = load_bottocen_from_file()


# Функции для работы с Excel

async def send_user_list():
    try:
        excel_filename = await export_to_excel()
        if excel_filename:
            bot = Bot(token=BOTTOCEN)
            with open(excel_filename, "rb") as file:
                filename_to_send = os.path.basename(excel_filename)
                await bot.send_document(
                    chat_id=CREATOR_CHAT_ID,
                    document=file,
                    filename=filename_to_send  # Явно указываем имя файла для отправки
                )
            # Удаляем временный файл
            try:
                os.remove(excel_filename)
            except Exception as e:
                print(f"Ошибка при удалении файла: {e}")
    except Exception as e:
        print(f"Error in send_user_list: {e}")
        try:
            bot = Bot(token=BOTTOCEN)
            await bot.send_message(chat_id=CREATOR_CHAT_ID, text=f"Помилка при створенні звіту: {e}")
        except:
            pass


async def export_to_excel():
    data = safe_json_read(DATA_FILE)

    # Генерация имени файла с безопасными символами
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Новый формат без двоеточий
    excel_filename = f"SupportBot_{current_time}.xlsx"  # Исправленное имя файла

    try:
        # Подготовка данных
        all_users_df = pd.DataFrame(data["users"])
        banned_ids = set(data["banned_users"].keys())

        # Переименовываем колонки для AllUsers
        all_users_df = all_users_df.rename(columns={
            'mute': 'mute/ban',
            'mute_end': 'mute/ban_end'
        })

        # Обновляем статус для забаненных пользователей
        for user_id in banned_ids:
            mask = all_users_df['id'] == user_id
            all_users_df.loc[mask, 'mute/ban'] = True
            all_users_df.loc[mask, 'mute/ban_end'] = "Навсегда (бан)"

        # Разделяем пользователей
        users_df = all_users_df[all_users_df["mute/ban"] == False].copy()
        muted_df = all_users_df[(all_users_df["mute/ban"] == True) & (~all_users_df['id'].isin(banned_ids))].copy()

        # Создаем отдельный лист для забаненных
        banned_users_info = []
        for user_id, ban_info in data["banned_users"].items():
            user_data = all_users_df[all_users_df["id"] == user_id].iloc[0].to_dict() if not all_users_df[
                all_users_df["id"] == user_id].empty else {
                "username": "Unknown",
                "first_name": "Unknown",
                "join_date": "",
                "rating": 0,
                "mute/ban": True,
                "mute/ban_end": "Навсегда (бан)"
            }

            banned_users_info.append({
                "id": user_id,
                "username": user_data.get("username", "Unknown"),
                "first_name": user_data.get("first_name", "Unknown"),
                "join_date": user_data.get("join_date", ""),
                "rating": user_data.get("rating", 0),
                "mute/ban": True,
                "mute/ban_end": "Навсегда (бан)",
                "reason": ban_info.get("reason", "Banned")
            })

        banned_df = pd.DataFrame(banned_users_info)

        # Создаем лист для topics и user_topics
        topics_df = pd.DataFrame({
            "user_id": list(data.get("topics", {}).keys()),
            "topic_id": list(data.get("topics", {}).values())
        })

        user_topics_df = pd.DataFrame({
            "topic_id": list(data.get("user_topics", {}).keys()),
            "user_id": list(data.get("user_topics", {}).values())
        })

        # Создаем лист для sent_messages
        sent_messages_df = pd.DataFrame([
            {"message_id": k, "user_id": v}
            for k, v in data.get("sent_messages", {}).items()
        ])

        # Обработка дат
        def process_dates(df):
            if "mute/ban_end" in df.columns:
                df.loc[:, "mute/ban_end"] = df["mute/ban_end"].apply(
                    lambda x: x if "Навсегда" in str(x) else
                    (datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y")
                     if isinstance(x, str) and x != "Навсегда" else "")
                )
            if "join_date" in df.columns:
                df.loc[:, "join_date"] = df["join_date"].apply(
                    lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime("%H:%M; %d/%m/%Y")
                    if isinstance(x, str) else ""
                )
            return df

        all_users_df = process_dates(all_users_df)
        users_df = process_dates(users_df)
        muted_df = process_dates(muted_df)
        banned_df = process_dates(banned_df)

        # Запись в Excel
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # Основные листы пользователей
            all_users_df.to_excel(writer, index=False, sheet_name="AllUsers")
            users_df.to_excel(writer, index=False, sheet_name="ActiveUsers")
            muted_df.to_excel(writer, index=False, sheet_name="MutedUsers")
            banned_df.to_excel(writer, index=False, sheet_name="BannedUsers")

            # Технические листы
            topics_df.to_excel(writer, index=False, sheet_name="Topics")
            user_topics_df.to_excel(writer, index=False, sheet_name="UserTopics")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")

            # Дополнительные листы
            pd.DataFrame(data.get("admins", []), columns=["Admins"]).to_excel(
                writer, index=False, sheet_name="Admins")
            pd.DataFrame(data.get("programmers", []), columns=["Programmers"]).to_excel(
                writer, index=False, sheet_name="Programmers")
            pd.DataFrame([{
                "bot_token": data.get("bot_token", ""),
                "owner_id": data.get("owner_id", ""),
                "chat_id": data.get("chat_id", ""),
                "total_score": data.get("total_score", 0),
                "num_of_ratings": data.get("num_of_ratings", 0)
            }]).to_excel(writer, index=False, sheet_name="GeneralInfo")

            # Стилизация
            workbook = writer.book

            # Цвета заливки
            light_blue_fill = PatternFill(start_color="8bbef2", end_color="8bbef2", fill_type="solid")
            light_green_fill = PatternFill(start_color="8bf28b", end_color="8bf28b", fill_type="solid")
            light_red_fill = PatternFill(start_color="f28b8b", end_color="f28b8b", fill_type="solid")
            light_yellow_fill = PatternFill(start_color="f2f28b", end_color="f2f28b", fill_type="solid")

            # Списки админов и программистов
            admins = data.get("admins", [])
            programmers = data.get("programmers", [])

            # Раскраска AllUsers
            if "AllUsers" in workbook.sheetnames:
                ws = workbook["AllUsers"]
                for row in ws.iter_rows(min_row=2):
                    user_id = row[0].value
                    username = row[1].value if len(row) > 1 else ""

                    if str(user_id) in banned_ids:
                        for cell in row:
                            cell.fill = light_red_fill
                    elif ws.cell(row=row[0].row, column=6).value == True:  # Колонка mute/ban
                        for cell in row:
                            cell.fill = light_yellow_fill
                    elif username in programmers:
                        for cell in row:
                            cell.fill = light_green_fill
                    elif username in admins:
                        for cell in row:
                            cell.fill = light_blue_fill

            # Раскраска ActiveUsers
            if "ActiveUsers" in workbook.sheetnames:
                ws = workbook["ActiveUsers"]
                for row in ws.iter_rows(min_row=2):
                    username = row[1].value if len(row) > 1 else ""

                    if username in programmers:
                        for cell in row:
                            cell.fill = light_green_fill
                    elif username in admins:
                        for cell in row:
                            cell.fill = light_blue_fill

            # Раскраска MutedUsers (только замученные)
            if "MutedUsers" in workbook.sheetnames:
                ws = workbook["MutedUsers"]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = light_yellow_fill

            # Раскраска BannedUsers (только забаненные)
            if "BannedUsers" in workbook.sheetnames:
                ws = workbook["BannedUsers"]
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = light_red_fill

        return excel_filename

    except Exception as e:
        logging.error(f"Ошибка при экспорте в Excel: {e}")
        return None

async def import_from_excel(file_path):
    try:
        data = safe_json_read(DATA_FILE)
        new_data = {
            "users": [],
            "muted_users": {},
            "banned_users": {},
            "admins": data.get("admins", []),
            "programmers": data.get("programmers", []),
            "bot_token": data.get("bot_token", ""),
            "owner_id": data.get("owner_id", ""),
            "chat_id": data.get("chat_id", ""),
            "total_score": data.get("total_score", 0),
            "num_of_ratings": data.get("num_of_ratings", 0),
            "sent_messages": {},
            "topics": {},
            "user_topics": {}
        }

        wb = load_workbook(file_path)

        # Импорт основных настроек из GeneralInfo
        if "GeneralInfo" in wb.sheetnames:
            ws = wb["GeneralInfo"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 5:
                    if len(headers) >= 1 and row[0]:  # bot_token
                        new_data["bot_token"] = str(row[0])
                    if len(headers) >= 2 and row[1]:  # owner_id
                        new_data["owner_id"] = str(row[1])
                    if len(headers) >= 3 and row[2]:  # chat_id
                        new_data["chat_id"] = str(row[2])
                    if len(headers) >= 4 and row[3] is not None:
                        new_data["total_score"] = float(row[3])
                    if len(headers) >= 5 and row[4] is not None:
                        new_data["num_of_ratings"] = int(row[4])

        # Сначала импортируем забаненных пользователей
        if "BannedUsers" in wb.sheetnames:
            ws = wb["BannedUsers"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3 and len(headers) >= 3:
                    user_id = str(row[0])
                    reason = row[headers.index("reason")] if "reason" in headers else "Импортирован из файла"

                    new_data["banned_users"][user_id] = {
                        "reason": reason,
                        "date": get_current_time_kiev()
                    }

        # Затем импортируем всех пользователей и заполняем muted_users
        if "AllUsers" in wb.sheetnames:
            ws = wb["AllUsers"]
            headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 7 and len(headers) >= 7:
                    user_data = dict(zip(headers[:7], row[:7]))

                    # Обработка mute/ban статуса
                    if "mute/ban" in user_data:
                        user_data["mute"] = user_data.pop("mute/ban")

                    # Обработка mute/ban_end
                    if "mute/ban_end" in user_data:
                        user_data["mute_end"] = user_data.pop("mute/ban_end")
                        if "Навсегда (бан)" in str(user_data["mute_end"]):
                            user_data["mute_end"] = "Навсегда"

                    # Добавляем в muted_users если пользователь замучен и не забанен
                    if user_data.get("mute", False) and user_data["id"] not in new_data["banned_users"]:
                        reason = user_data.get("reason", "Причина не указана")  # Всегда сохраняем причину
                        new_data["muted_users"][user_data["id"]] = {
                            "expiration": user_data.get("mute_end"),
                            "reason": reason if reason else "Причина не указана"  # Гарантируем наличие причины
                        }

                    new_data["users"].append(user_data)

        # Импорт дополнительных данных
        if "Topics" in wb.sheetnames:
            ws = wb["Topics"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["topics"][str(row[0])] = row[1]

        if "UserTopics" in wb.sheetnames:
            ws = wb["UserTopics"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["user_topics"][str(row[0])] = str(row[1])

        if "SentMessages" in wb.sheetnames:
            ws = wb["SentMessages"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2:
                    new_data["sent_messages"][str(row[0])] = str(row[1])

        if "Admins" in wb.sheetnames:
            ws = wb["Admins"]
            new_data["admins"] = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]

        if "Programmers" in wb.sheetnames:
            ws = wb["Programmers"]
            new_data["programmers"] = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]

        safe_json_write(new_data, DATA_FILE)
        return True

    except Exception as e:
        print(f"Ошибка при импорте из Excel: {e}")
        return False

async def auto_delete_message(bot, chat_id, message_id, delay):
    await asyncio.sleep(delay)
    try:
        await bot.delete_message(chat_id=chat_id, message_id=message_id)
    except:
        pass


async def start(update: Update, context):
    try:
        user = update.message.from_user
        chat_id = update.effective_chat.id

        if chat_id == CREATOR_CHAT_ID:
            await update.message.reply_text("Команда /start недоступна в цій групі.")
            return

        config = safe_json_read(DATA_FILE)
        user_found = False

        for u in config["users"]:
            if u["id"] == str(user.id):
                user_found = True
                break

        if not user_found:
            new_user = {
                "id": str(user.id),
                "username": user.username or "Не вказано",
                "first_name": user.first_name or "Не вказано",
                "join_date": get_current_time_kiev(),
                "rating": 0,
                "mute": False,
                "mute_end": None,
                "reason": None
            }
            config["users"].append(new_user)
            safe_json_write(config, DATA_FILE)

        keyboard = [
            ["/start", "/rate"],
            ["/message", "/stopmessage"],
            ["/fromus", "/help"],
        ]

        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

        await update.message.reply_text(
            "Привіт! Я ваш бот підтримки. Введіть команду /rate для оцінки бота, /message для написання адміністраторам або /help для допомоги.",
            reply_markup=reply_markup
        )
    except Exception as e:
        print(f"Error in start: {e}")
        await update.message.reply_text("Сталася помилка. Спробуйте ще раз.")


async def rate(update: Update, context):
    try:
        user_id = update.message.from_user.id
        data = safe_json_read(DATA_FILE)

        user_rating = None
        for user in data.get("users", []):
            if user.get('id') == str(user_id):
                user_rating = user['rating']
                break

        total_score = data.get("total_score", 0)
        num_of_ratings = data.get("num_of_ratings", 0)
        average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

        rating_text = f"Загальна оцінка: {round(average_rating, 1)}⭐️\nВаш попередній відгук: {user_rating}⭐️" if user_rating else f"Загальна оцінка: {round(average_rating, 1)}⭐️"

        keyboard = [
            [InlineKeyboardButton("0.5⭐️", callback_data='0.5'), InlineKeyboardButton("1⭐️", callback_data='1')],
            [InlineKeyboardButton("1.5⭐️", callback_data='1.5'), InlineKeyboardButton("2⭐️", callback_data='2')],
            [InlineKeyboardButton("2.5⭐️", callback_data='2.5'), InlineKeyboardButton("3⭐️", callback_data='3')],
            [InlineKeyboardButton("3.5⭐️", callback_data='3.5'), InlineKeyboardButton("4⭐️", callback_data='4')],
            [InlineKeyboardButton("4.5⭐️", callback_data='4.5'), InlineKeyboardButton("5⭐️", callback_data='5')],
        ]

        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(f"{rating_text}\nОберіть оцінку:", reply_markup=reply_markup)
    except Exception as e:
        print(f"Error in rate: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def button_callback(update: Update, context):
    try:
        query = update.callback_query
        await query.answer()
        user_id = query.from_user.id
        new_rating = float(query.data)

        data = safe_json_read(DATA_FILE)
        user_found = False
        previous_rating = 0

        for user in data.get("users", []):
            if user.get('id') == str(user_id):
                previous_rating = user.get('rating', 0)
                user['rating'] = new_rating
                user_found = True
                break

        if not user_found:
            new_user = {
                'id': str(user_id),
                'first_name': query.from_user.first_name,
                'username': query.from_user.username,
                'join_date': get_current_time_kiev(),
                'rating': new_rating,
                'mute': False,
                'mute_end': None,
                'reason': None
            }
            data['users'].append(new_user)

        total_score = data.get("total_score", 0)
        num_of_ratings = data.get("num_of_ratings", 0)

        if previous_rating == 0:
            num_of_ratings += 1
            total_score += new_rating
        else:
            total_score = total_score - previous_rating + new_rating

        data["total_score"] = total_score
        data["num_of_ratings"] = num_of_ratings

        safe_json_write(data, DATA_FILE)

        average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0
        await query.edit_message_text(
            f"Дякуємо за ваш відгук! Ваша оцінка: {new_rating}⭐️\nЗагальна оцінка: {round(average_rating, 1)}⭐️"
        )
    except Exception as e:
        print(f"Error in button_callback: {e}")
        await query.edit_message_text("Сталася помилка при обробці вашого відгуку.")


async def message(update: Update, context):
    try:
        muted_users = load_muted_users_from_file()
        user_id = update.message.from_user.id

        if str(user_id) in muted_users:
            mute_info = muted_users[str(user_id)]
            if mute_info['expiration'] and mute_info['expiration'] > datetime.now():
                reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        context.user_data['waiting_for_message'] = True
        reply = await update.message.reply_text(
            "Введіть ваше повідомлення, і його буде відправлено адміністраторам бота. "
            "Введіть /stopmessage, щоб завершити введення повідомлень."
        )
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    except Exception as e:
        print(f"Error in message command: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def stopmessage(update: Update, context):
    try:
        if context.user_data.get('waiting_for_message'):
            reply = await update.message.reply_text("Ви завершили введення повідомлень.")
            context.user_data['waiting_for_message'] = False
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
        else:
            await update.message.reply_text("Ви не в режимі введення повідомлень.")
    except Exception as e:
        print(f"Error in stopmessage: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def help(update: Update, context):
    try:
        if str(update.message.chat.id) == str(CREATOR_CHAT_ID):
            help_text = (
                "Доступні команди в групі:\n"
                "Відповісти на повідомлення бота - Надіслати повідомлення користувачу, який надіслав це повідомлення.\n"
                "/mute <час> <користувач> 'причина' - Замутити користувача на вказаний час.\n"
                "/unmute <користувач> - Розмутити користувача.\n"
                "/mutelist - Показати список замучених користувачів.\n"
                "/alllist - Показати всіх користувачів.\n"
                "/allmessage <повідомлення> - Надіслати повідомлення всім користувачам.\n"
                "/fromus - Інформація про створювача.\n"
                "/help - Показати доступні команди.\n"
                "/info - Показати інформацію про програмістів та адміністраторів.\n"
                "/admin <користувач> - Додати адміністратора.\n"
                "/deleteadmin <користувач> - Видалити адміністратора.\n"
                "/programier <користувач> - Додати програміста.\n"
                "/deleteprogramier <користувач> - Видалити програміста.\n"
                "/get_alllist - Отримати Excel файл з користувачами.\n"
                "/set_alllist - Записати Excel файл з користувачами.\n"
            )
        elif str(update.message.chat.id) == str(-1002358066044):
            help_text = (
                "Доступні команди в групі:\n"
                "/get_alllist - Отримати Excel файл з користувачами.\n"
                "/set_alllist - Записати Excel файл з користувачами.\n"
            )
        else:
            help_text = (
                "Доступні команди в боті:\n"
                "/start - Запустити бота.\n"
                "/rate - Залишити відгук.\n"
                "/message - Почати введення повідомлень адміністраторам.\n"
                "/stopmessage - Завершити введення повідомлень.\n"
                "/fromus - Інформація про створювача.\n"
                "/help - Показати доступні команди.\n"
            )

        await update.message.reply_text(help_text)
    except Exception as e:
        print(f"Error in help: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def fromus(update: Update, context):
    try:
        await update.message.reply_text(
            "*Skeleton* Написав бота\nПортфоліо: ```https://www.linkedin.com/in/artem-k-972a41344/```\n"
            "Телеграм канал з усіма проєктами: ```https://t.me/AboutMyProjects```\n"
            "По всім питанням пишіть в цього бота",
            parse_mode="MarkdownV2"
        )
    except Exception as e:
        print(f"Error in fromus: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def info(update: Update, context: CallbackContext):
    try:
        data = safe_json_read(DATA_FILE)
        programmers = data.get("programmers", [])
        admins = data.get("admins", [])

        programmer_list = "\n".join(programmers) if programmers else "Список програмістів пустий."
        admin_list = "\n".join(admins) if admins else "Список адміністраторів пустий."

        await update.message.reply_text(f"Програмісти:\n{programmer_list}\n\nАдміністратори:\n{admin_list}")
    except Exception as e:
        print(f"Error in info: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def mute(update: Update, context: CallbackContext):
    try:
        # Проверка прав администратора
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Эта команда доступна только администраторам.")
            return

        # Получаем topic_id из темы
        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Эта команда работает только в темах пользователей.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не удалось определить пользователя для этой темы.")
            return

        # Проверяем, не забанен ли пользователь
        if user_id in data["banned_users"]:
            await update.message.reply_text("❌ Этот пользователь забанен и не может быть замучен!")
            return

        # Парсим время и причину
        mute_time = 300  # 5 минут по умолчанию
        reason = "По решению администратора"

        if context.args:
            if context.args[0].isdigit():
                mute_time = int(context.args[0])
                if len(context.args) > 1:
                    reason = ' '.join(context.args[1:])
            else:
                reason = ' '.join(context.args)

        # Проверяем пользователя
        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Пользователь не найден.")
            return

        if user_data["id"] == data["owner_id"]:
            await update.message.reply_text("Невозможно замутить владельца чата.")
            return

        # Устанавливаем мут
        mute_end = (datetime.now() + timedelta(seconds=mute_time)).strftime("%H:%M; %d/%m/%Y")
        user_data.update({
            "mute": True,
            "mute_end": mute_end,
            "reason": reason
        })

        data["muted_users"][user_id] = {
            "expiration": mute_end,
            "reason": reason
        }

        safe_json_write(data, DATA_FILE)

        # Новый формат ChatPermissions
        mute_permissions = ChatPermissions(
            can_send_messages=False,
            can_send_photos=False,
            can_send_videos=False,
            can_send_audios=False,
            can_send_documents=False,
            can_send_polls=False,
            can_send_other_messages=False,
            can_add_web_page_previews=False,
            can_change_info=False,
            can_invite_users=False,
            can_pin_messages=False
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=mute_permissions,
            until_date=int((datetime.now() + timedelta(seconds=mute_time)).timestamp())
        )

        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"🔇 Вас замутили на {mute_time} секунд\nПричина: {reason}\nМут истечет: {mute_end}"
            )
        except Exception as e:
            print(f"Ошибка уведомления о муте: {e}")

        await update.message.reply_text(f"✅ Пользователь замучен на {mute_time} секунд. Причина: {reason}")

    except Exception as e:
        print(f"Ошибка в команде mute: {e}")
        await update.message.reply_text("❌ Произошла ошибка при обработке команды.")

async def unmute(update: Update, context: CallbackContext):
    try:
        # Проверка прав администратора
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Эта команда доступна только администраторам.")
            return

        # Получаем topic_id из темы
        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Эта команда работает только в темах пользователей.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не удалось определить пользователя для этой темы.")
            return

        # Проверяем, не забанен ли пользователь
        if user_id in data["banned_users"]:
            await update.message.reply_text("❌ Этот пользователь забанен! Используйте /unban для разбана.")
            return

        # Проверяем статус пользователя
        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Пользователь не найден.")
            return

        if not user_data["mute"]:
            await update.message.reply_text("Этот пользователь не в муте.")
            return

        # Снимаем мут
        user_data.update({
            "mute": False,
            "mute_end": None,
            "reason": None
        })

        if user_id in data["muted_users"]:
            del data["muted_users"][user_id]

        safe_json_write(data, DATA_FILE)

        # Новый формат ChatPermissions
        unmute_permissions = ChatPermissions(
            can_send_messages=True,
            can_send_photos=True,
            can_send_videos=True,
            can_send_audios=True,
            can_send_documents=True,
            can_send_polls=True,
            can_send_other_messages=True,
            can_add_web_page_previews=True,
            can_change_info=True,
            can_invite_users=True,
            can_pin_messages=True
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=unmute_permissions
        )

        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text="🔊 Вас размутили. Теперь вы можете снова писать в чат."
            )
        except Exception as e:
            print(f"Ошибка уведомления о размуте: {e}")

        await update.message.reply_text(f"✅ Пользователь @{user_data['username']} был размучен.")

    except Exception as e:
        print(f"Ошибка в команде unmute: {e}")
        await update.message.reply_text("❌ Произошла ошибка при обработке команды.")

async def ban(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Эта команда доступна только администраторам.")
            return

        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Эта команда работает только в темах пользователей.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не удалось определить пользователя для этой темы.")
            return

        reason = "По решению администратора"
        if context.args:
            reason = ' '.join(context.args)

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if not user_data:
            await update.message.reply_text("Пользователь не найден.")
            return

        if user_data["id"] == data["owner_id"]:
            await update.message.reply_text("Невозможно забанить владельца чата.")
            return

        data["banned_users"][user_id] = {
            "reason": reason,
            "date": get_current_time_kiev()
        }

        user_data.update({
            "mute": True,
            "mute_end": "Навсегда",
            "reason": f"Забанен: {reason}"
        })

        data["muted_users"][user_id] = {
            "expiration": "Навсегда",
            "reason": f"Забанен: {reason}"
        }

        safe_json_write(data, DATA_FILE)

        await context.bot.ban_chat_member(
            chat_id=data["chat_id"],
            user_id=int(user_id)
        )

        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=f"🚫 Вас забанили навсегда\nПричина: {reason}"
            )
        except Exception as e:
            print(f"Ошибка уведомления пользователя о бане: {e}")

        await update.message.reply_text(f"Пользователь забанен навсегда. Причина: {reason}")

    except Exception as e:
        print(f"Ошибка в команде ban: {e}")
        await update.message.reply_text("Произошла ошибка при обработке команды.")


async def unban(update: Update, context: CallbackContext):
    try:
        # Проверка прав администратора
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Эта команда доступна только администраторам.")
            return

        # Получаем topic_id из темы
        topic_id = update.message.message_thread_id
        if not topic_id:
            await update.message.reply_text("Эта команда работает только в темах пользователей.")
            return

        data = safe_json_read(DATA_FILE)
        user_id = data.get("user_topics", {}).get(str(topic_id))
        if not user_id:
            await update.message.reply_text("Не удалось определить пользователя для этой темы.")
            return

        if user_id not in data["banned_users"]:
            await update.message.reply_text("Этот пользователь не забанен.")
            return

        # Снимаем бан
        del data["banned_users"][user_id]

        user_data = next((u for u in data["users"] if u["id"] == user_id), None)
        if user_data:
            user_data.update({
                "mute": False,
                "mute_end": None,
                "reason": None
            })

        if user_id in data["muted_users"]:
            del data["muted_users"][user_id]

        safe_json_write(data, DATA_FILE)

        # Разбаниваем
        await context.bot.unban_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id)
        )

        # Новый формат ChatPermissions
        unmute_permissions = ChatPermissions(
            can_send_messages=True,
            can_send_photos=True,
            can_send_videos=True,
            can_send_audios=True,
            can_send_documents=True,
            can_send_polls=True,
            can_send_other_messages=True,
            can_add_web_page_previews=True,
            can_change_info=True,
            can_invite_users=True,
            can_pin_messages=True
        )

        await context.bot.restrict_chat_member(
            chat_id=int(data["chat_id"]),
            user_id=int(user_id),
            permissions=unmute_permissions
        )

        # Уведомляем пользователя
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text="✅ Вас разбанили. Теперь вы можете снова участвовать в чате."
            )
        except Exception as e:
            print(f"Ошибка уведомления о разбане: {e}")

        await update.message.reply_text(f"✅ Пользователь разбанен.")

    except Exception as e:
        print(f"Ошибка в команде unban: {e}")
        await update.message.reply_text("❌ Произошла ошибка при обработке команды.")


async def check_mute_expirations():
    try:
        application = Application.builder().token(BOTTOCEN).build()
        async with application:
            context = ContextTypes.DEFAULT_TYPE(application=application)

            data = safe_json_read(DATA_FILE)
            now = datetime.now()
            users_to_unmute = []

            for user in data["users"]:
                if user["mute"] and user["mute_end"]:
                    try:
                        # Преобразуем строку времени в объект datetime
                        mute_end = datetime.strptime(user["mute_end"], "%H:%M; %d/%m/%Y")
                        if mute_end <= now:
                            users_to_unmute.append(user)
                    except ValueError:
                        continue

            if users_to_unmute:
                for user in users_to_unmute:
                    # Обновляем данные пользователя
                    user.update({
                        "mute": False,
                        "mute_end": None,
                        "reason": None
                    })

                    # Удаляем из muted_users если есть
                    if "muted_users" in data and user["id"] in data["muted_users"]:
                        del data["muted_users"][user["id"]]

                    # Снимаем ограничения
                    try:
                        await context.bot.restrict_chat_member(
                            chat_id=int(data["chat_id"]),
                            user_id=int(user["id"]),
                            permissions=ChatPermissions(
                                can_send_messages=True,
                                can_send_photos=True,
                                can_send_videos=True,
                                can_send_audios=True,
                                can_send_documents=True,
                                can_send_polls=True,
                                can_send_other_messages=True,
                                can_add_web_page_previews=True,
                                can_change_info=True,
                                can_invite_users=True,
                                can_pin_messages=True
                            )
                        )

                        # Уведомляем пользователя
                        try:
                            await context.bot.send_message(
                                chat_id=int(user["id"]),
                                text="🔊 Ваш мут истек. Теперь вы можете снова писать в чат."
                            )
                        except Exception as e:
                            print(f"Ошибка уведомления пользователя об окончании мута: {e}")

                    except Exception as e:
                        print(f"Ошибка при размуте пользователя {user['id']}: {e}")

                # Сохраняем изменения в файл
                print(f"Данные перед сохранением: {data['users']}")
                if not safe_json_write(data, DATA_FILE):
                    print("Ошибка сохранения данных")
                else:
                    print("Данные успешно сохранены")
                    # Проверяем, что сохранилось
                    check_data = safe_json_read(DATA_FILE)
                    print(f"Данные после сохранения: {check_data['users']}")

    except Exception as e:
        print(f"Ошибка в проверке сроков мута: {e}")

async def admin(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /admin @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["admins"]:
            await update.message.reply_text(f"Користувач @{username} вже є адміністратором.")
        else:
            data["admins"].append(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"Користувач @{username} доданий до списку адміністраторів.")
    except Exception as e:
        print(f"Error in admin: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def deleteadmin(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /deleteadmin @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["admins"]:
            data["admins"].remove(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"Користувач @{username} видалений зі списку адміністраторів.")
        else:
            await update.message.reply_text(f"Користувач @{username} не знайдений.")
    except Exception as e:
        print(f"Error in deleteadmin: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def programier(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /programier @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username in data["programmers"]:
            await update.message.reply_text(f"Користувач @{username} вже є програмістом.")
        else:
            data["programmers"].append(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"Користувач @{username} доданий до списку програмістів.")
    except Exception as e:
        print(f"Error in programier: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def deleteprogramier(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user):
            await update.message.reply_text("Ця команда доступна тільки програмістам.")
            return

        if len(context.args) < 1:
            await update.message.reply_text("Використовуйте: /deleteprogramier @username")
            return

        username = context.args[0].lstrip('@')
        data = safe_json_read(DATA_FILE)

        if username == "ArtemKirss":
            await update.message.reply_text(f"Неможливо видалити {username} зі списку програмістів.")
        elif username in data["programmers"]:
            data["programmers"].remove(username)
            safe_json_write(data, DATA_FILE)
            await update.message.reply_text(f"Користувач @{username} видалений зі списку програмістів.")
        else:
            await update.message.reply_text(f"Користувач @{username} не є програмістом.")
    except Exception as e:
        print(f"Error in deleteprogramier: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def mutelist(update: Update, context):
    try:
        user = update.message.from_user.username
        if update.message.chat.id != CREATOR_CHAT_ID:
            if not is_programmer(user) and not is_admin(user):
                reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        data = safe_json_read(DATA_FILE)
        admins = data.get("admins", [])
        programmers = data.get("programmers", [])
        muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

        response = "Замучені користувачі:\n"

        if muted_users:
            for user_id, mute_info in muted_users.items():
                expiration = mute_info.get('mute_end', 'Невідомо')
                reason = mute_info.get('reason', 'Без причини')

                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"

                join_date = mute_info.get('join_date', 'Невідома')
                rating = mute_info.get('rating', 0)
                mute_symbol = "🔇"

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                response += (
                    f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                    f"Залишилось: {expiration}\n"
                    f"Причина: {reason}\n"
                    f"Дата заходу: {join_date}\n"
                    f"Оцінка: {rating}⭐️\n"
                    "-------------------------------------------------------------------------\n"
                )
        else:
            response += "Немає замучених користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        await update.message.reply_text(response)
    except Exception as e:
        print(f"Error in mutelist: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def alllist(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if update.message.chat.id != CREATOR_CHAT_ID:
            if not is_programmer(user) and not is_admin(user):
                reply = await update.message.reply_text("Ця команда доступна лише адміністраторам бота.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        data = safe_json_read(DATA_FILE)
        admins = data.get("admins", [])
        programmers = data.get("programmers", [])
        users_info = {user['id']: user for user in data.get("users", [])}
        muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

        response = "Користувачі:\n"
        unique_users = {user['id'] for user in data.get("users", [])}

        if unique_users:
            for user_id in unique_users:
                user_data = users_info.get(str(user_id), {})
                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"
                join_date = user_data.get('join_date', 'Невідома')
                rating = user_data.get('rating', 0)

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                mute_symbol = "🔇" if str(user_id) in muted_users else "🔊"

                response += f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\nДата заходу: {join_date}\nОцінка: {rating}⭐️\n"
                response += "-------------------------------------------------------------------------\n"
        else:
            response += "Немає користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        response += "==========================================\n"
        response += "\n"
        response += "==========================================\n"
        response += "Замучені користувачі:\n"

        if muted_users:
            for user_id, mute_info in muted_users.items():
                expiration = mute_info['mute_end'] or "Невідомо"
                reason = mute_info.get('reason', "Без причини")
                user_info = await context.bot.get_chat_member(chat_id=data["chat_id"], user_id=int(user_id))
                user_fullname = user_info.user.first_name or "Невідомий"
                username = user_info.user.username or "Немає імені користувача"
                user_data = users_info.get(str(user_id), {})
                join_date = user_data.get('join_date', 'Невідома')
                rating = user_data.get('rating', 0)

                admins_sumdol = "👨🏻‍💼"
                if username in admins:
                    admins_sumdol = "👮🏻‍♂️"
                if username in programmers:
                    admins_sumdol = "👨🏻‍💻"

                mute_symbol = "🔇"

                response += (
                    f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                    f"Залишилось: {str(expiration).split('.')[0]}\n"
                    f"Причина: {reason}\n"
                    f"Дата заходу: {join_date}\n"
                    f"Оцінка: {rating}⭐️\n"
                    "-------------------------------------------------------------------------\n"
                )
        else:
            response += "Немає замучених користувачів.\n"
            response += "-------------------------------------------------------------------------\n"

        await update.message.reply_text(response)
    except Exception as e:
        print(f"Error in alllist: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")


async def allmessage(update: Update, context):
    try:
        user = update.message.from_user.username

        if update.message.chat.id != CREATOR_CHAT_ID:
            if not is_programmer(user) and not is_admin(user):
                reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

        if not context.args:
            await update.message.reply_text("Будь ласка, укажіть текст повідомлення після команди.")
            return

        message_text = update.message.text.split(' ', 1)[1]
        data = safe_json_read(DATA_FILE)

        for user_data in data.get("users", []):
            user_id = user_data.get("id")
            if user_id:
                try:
                    await context.bot.send_message(chat_id=user_id, text=message_text)
                except Exception as e:
                    print(f"Помилка при відправці повідомлення користувачу {user_id}: {e}")

        await update.message.reply_text("Повідомлення відправлено всім користувачам.")
    except Exception as e:
        print(f"Error in allmessage: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди.")

async def get_alllist(update: Update, context: CallbackContext) -> None:
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
            return

        excel_filename = await export_to_excel()
        if excel_filename:
            # Читаем файл в бинарном режиме
            with open(excel_filename, "rb") as file:
                # Создаем имя файла для отправки (без пути)
                filename_to_send = os.path.basename(excel_filename)
                await update.message.reply_document(
                    document=file,
                    filename=filename_to_send  # Явно указываем имя файла для отправки
                )
            # Удаляем временный файл
            try:
                os.remove(excel_filename)
            except Exception as e:
                print(f"Ошибка при удалении файла: {e}")
        else:
            await update.message.reply_text("Помилка при створенні Excel-файлу")
    except Exception as e:
        print(f"Error in get_alllist: {e}")
        await update.message.reply_text("Сталася помилка при експорті даних")


async def set_alllist(update: Update, context: CallbackContext) -> None:
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
            return

        await update.message.reply_text("Будь ласка, надішліть Excel-файл з даними.")
        context.user_data["awaiting_file"] = True
    except Exception as e:
        print(f"Error in set_alllist: {e}")
        await update.message.reply_text("Сталася помилка")


async def get_logs(update: Update, context: CallbackContext):
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Эта команда доступна только администраторам.")
            return

        log_file = "bot_errors.log"

        if not os.path.exists(log_file):
            await update.message.reply_text("Файл логов не найден.")
            return

        with open(log_file, "rb") as file:
            await update.message.reply_document(
                document=file,
                filename=f"bot_logs_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
            )

    except Exception as e:
        print(f"Ошибка при отправке логов: {e}")
        await update.message.reply_text("❌ Произошла ошибка при отправке логов.")

async def write_to_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        user = update.message.from_user.username
        if not is_programmer(user) and not is_admin(user):
            await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
            return

        if len(context.args) < 2:
            await update.message.reply_text("Використовуйте: /write <user_id> <текст повідомлення>")
            return

        user_id = context.args[0]
        message_text = ' '.join(context.args[1:])
        data = safe_json_read(DATA_FILE)

        # Получаем topic_id для этого пользователя
        topic_id = data.get("topics", {}).get(str(user_id))

        if topic_id:
            # Отправляем сообщение в тему
            await context.bot.send_message(
                chat_id=data["chat_id"],
                message_thread_id=topic_id,
                text=message_text
            )
            await update.message.reply_text("Повідомлення відправлено в тему користувача")
        else:
            await update.message.reply_text("Тема для цього користувача не знайдена")
    except Exception as e:
        print(f"Error in write_to_user: {e}")
        await update.message.reply_text("Сталася помилка при обробці команди")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        sent_messages = load_sent_messages()
        muted_users = load_muted_users_from_file()
        data = safe_json_read(DATA_FILE)

        # Обработка загрузки файла
        if context.user_data.get("awaiting_file"):
            if update.message.document:
                file = await update.message.document.get_file()
                await file.download_to_drive("temp_import.xlsx")

                if await import_from_excel("temp_import.xlsx"):
                    await update.message.reply_text("Дані успішно імпортовано!")
                else:
                    await update.message.reply_text("Помилка при імпорті даних")

                context.user_data["awaiting_file"] = False
                try:
                    os.remove("temp_import.xlsx")
                except:
                    pass
                return

        # Проверка на сообщение в главной теме (General)
        if update.message.chat.id == int(data["chat_id"]) and update.message.message_thread_id is None:
            user = update.message.from_user.username
            if is_programmer(user) or is_admin(user):
                # Отправляем сообщение всем пользователям
                success_count = 0
                fail_count = 0

                for user_data in data.get("users", []):
                    user_id = user_data.get("id")
                    if user_id:
                        try:
                            # Для текстовых сообщений
                            if update.message.text:
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text=f"📢 <b>Оголошення від адміністрації:</b>\n{update.message.text}",
                                    parse_mode='HTML'
                                )
                            # Для фото с подписью
                            elif update.message.photo:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_photo(
                                    chat_id=int(user_id),
                                    photo=update.message.photo[-1].file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                                # Обработка документов
                            elif update.message.document:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_document(
                                    chat_id=int(user_id),
                                    document=update.message.document.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                                # Обработка стикеров (НОВОЕ)
                            elif update.message.sticker:
                                # Сначала отправляем текст об объявлении
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text="📢 <b>Оголошення від адміністрації:</b>",
                                    parse_mode='HTML'
                                )
                                # Затем отправляем сам стикер
                                await context.bot.send_sticker(
                                    chat_id=int(user_id),
                                    sticker=update.message.sticker.file_id
                                )
                                # Обработка голосовых сообщений
                            elif update.message.voice:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_voice(
                                    chat_id=int(user_id),
                                    voice=update.message.voice.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                                # Обработка видео
                            elif update.message.video:
                                caption = f"📢 <b>Оголошення від адміністрації</b>"
                                if update.message.caption:
                                    caption += f":\n{update.message.caption}"
                                await context.bot.send_video(
                                    chat_id=int(user_id),
                                    video=update.message.video.file_id,
                                    caption=caption,
                                    parse_mode='HTML'
                                )
                                # Обработка видеосообщений (кружочки)
                            elif update.message.video_note:
                                await context.bot.send_message(
                                    chat_id=int(user_id),
                                    text="📢 <b>Оголошення від адміністрації:</b>",
                                    parse_mode='HTML'
                                )
                                await context.bot.send_video_note(
                                    chat_id=int(user_id),
                                    video_note=update.message.video_note.file_id
                                )

                            success_count += 1
                        except Exception as e:
                            print(f"Ошибка отправки сообщения пользователю {user_id}: {str(e)}")
                            fail_count += 1

                # Формируем отчет о рассылке
                report_message = (
                    f"📊 <b>Результат розсилки:</b>\n"
                    f"• Відправлено: {success_count}\n"
                    f"• Не вдалося: {fail_count}\n"
                    f"• Усього користувачів: {len(data.get('users', []))}"
                )

                await update.message.reply_text(
                    report_message,
                    parse_mode='HTML'
                )
                return

        # Обработка сообщений от пользователей в личном чате
        if str(update.message.chat.id) != str(data["chat_id"]):
            user_id = update.message.from_user.id
            if str(user_id) in muted_users and muted_users[str(user_id)]['expiration'] and muted_users[str(user_id)][
                'expiration'] > datetime.now():
                reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
                await asyncio.create_task(
                    auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
                return

            if context.user_data.get('waiting_for_message'):
                user_name = update.effective_user.first_name
                user_username = update.effective_user.username if update.effective_user.username else "немає імені користувача"
                current_time = get_current_time_kiev()

                # Экранируем специальные символы для MarkdownV2
                def escape_markdown(text):
                    if not text:
                        return ""
                    escape_chars = r'_*[]()~`>#+-=|{}.!'
                    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', text)

                # Получаем или создаем тему для пользователя
                topic_id = await get_or_create_topic(context, user_id, user_name)

                if topic_id:
                    # Формируем базовое сообщение
                    base_message = f'Повідомлення від **{escape_markdown(user_name)}**; `@{escape_markdown(user_username)}` `{user_id}`\n{escape_markdown(current_time)}:'

                    # Отправляем в зависимости от типа сообщения
                    if update.message.text:
                        message_text = f'{base_message}\n{escape_markdown(update.message.text)}'
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.photo:
                        photo_file_id = update.message.photo[-1].file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_photo(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            photo=photo_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.document:
                        document_file_id = update.message.document.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_document(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            document=document_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.sticker:
                        sticker_file_id = update.message.sticker.file_id
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=base_message,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        await context.bot.send_sticker(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            sticker=sticker_file_id
                        )
                        save_sent_messages(sent_messages)
                    elif update.message.voice:
                        voice_file_id = update.message.voice.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_voice(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            voice=voice_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.video:
                        video_file_id = update.message.video.file_id
                        caption = update.message.caption if update.message.caption else ''
                        message_text = f'{base_message}\n{escape_markdown(caption)}' if caption else base_message
                        msg = await context.bot.send_video(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            video=video_file_id,
                            caption=message_text,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        save_sent_messages(sent_messages)
                    elif update.message.video_note:
                        video_note_file_id = update.message.video_note.file_id
                        msg = await context.bot.send_message(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            text=base_message,
                            parse_mode="MarkdownV2"
                        )
                        sent_messages[str(msg.message_id)] = user_id
                        await context.bot.send_video_note(
                            chat_id=data["chat_id"],
                            message_thread_id=topic_id,
                            video_note=video_note_file_id
                        )
                        save_sent_messages(sent_messages)

                    reply = await update.message.reply_text("Ваше повідомлення надіслано адміністраторам бота.")
                    await asyncio.create_task(
                        auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
            else:
                await update.message.reply_text("Введіть /message, щоб надсилати повідомлення адміністраторам бота.")
            return

        # Обработка сообщений администраторов в темах форума
        if update.message.message_thread_id is not None:
            # Проверяем права администратора
            user = update.message.from_user.username
            if not is_programmer(user) and not is_admin(user):
                return

            # Получаем user_id из topic_id
            user_topics = data.get("user_topics", {})
            user_id = user_topics.get(str(update.message.message_thread_id))

            if user_id:
                try:
                    # Отправляем сообщение пользователю
                    if update.message.text:
                        await context.bot.send_message(
                            chat_id=int(user_id),
                            text=update.message.text
                        )
                    elif update.message.photo:
                        await context.bot.send_photo(
                            chat_id=int(user_id),
                            photo=update.message.photo[-1].file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.document:
                        await context.bot.send_document(
                            chat_id=int(user_id),
                            document=update.message.document.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.sticker:
                        await context.bot.send_sticker(
                            chat_id=int(user_id),
                            sticker=update.message.sticker.file_id
                        )
                    elif update.message.voice:
                        await context.bot.send_voice(
                            chat_id=int(user_id),
                            voice=update.message.voice.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video:
                        await context.bot.send_video(
                            chat_id=int(user_id),
                            video=update.message.video.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video_note:
                        await context.bot.send_video_note(
                            chat_id=int(user_id),
                            video_note=update.message.video_note.file_id
                        )

                    await update.message.reply_text("Повідомлення відправлено користувачу")
                except Exception as e:
                    await update.message.reply_text(f"Помилка при відправці: {str(e)}")
            return

        # Обработка ответов на сообщения бота (старая функциональность)
        if update.message.reply_to_message and update.message.reply_to_message.from_user.id == context.bot.id:
            original_message_id = str(update.message.reply_to_message.message_id)
            if original_message_id in sent_messages:
                original_user_id = sent_messages[original_message_id]
                reply_text = update.message.text if update.message.text else ""

                # Находим имя пользователя
                user_name = "Користувач"
                for user_data in data['users']:
                    if str(user_data['id']) == str(original_user_id):
                        user_name = user_data['first_name']
                        break

                # Отправляем сообщение пользователю
                try:
                    if update.message.photo:
                        await context.bot.send_photo(
                            chat_id=int(original_user_id),
                            photo=update.message.photo[-1].file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.document:
                        await context.bot.send_document(
                            chat_id=int(original_user_id),
                            document=update.message.document.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.sticker:
                        await context.bot.send_sticker(
                            chat_id=int(original_user_id),
                            sticker=update.message.sticker.file_id
                        )
                    elif update.message.voice:
                        await context.bot.send_voice(
                            chat_id=int(original_user_id),
                            voice=update.message.voice.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video:
                        await context.bot.send_video(
                            chat_id=int(original_user_id),
                            video=update.message.video.file_id,
                            caption=update.message.caption if update.message.caption else None
                        )
                    elif update.message.video_note:
                        await context.bot.send_video_note(
                            chat_id=int(original_user_id),
                            video_note=update.message.video_note.file_id
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=int(original_user_id),
                            text=reply_text
                        )

                    await update.message.reply_text(f"Користувачу {user_name} було надіслано повідомлення")
                    sent_messages[str(update.message.message_id)] = update.message.from_user.id
                    save_sent_messages(sent_messages)
                except Exception as e:
                    await update.message.reply_text(f"Помилка при відправці: {str(e)}")
    except Exception as e:
        print(f"Error in handle_message: {str(e)}")

def escape_markdown(text):
    if not text:
        return ""
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', text)

async def get_or_create_topic(context: ContextTypes.DEFAULT_TYPE, user_id: int, first_name: str):
    try:
        data = safe_json_read(DATA_FILE)
        chat_id = int(data["chat_id"])
        topics = data.get("topics", {})
        user_topics = data.get("user_topics", {})

        # Если тема уже существует, возвращаем её ID
        if str(user_id) in topics:
            return topics[str(user_id)]

        # Создаем новую тему
        topic_name = f"{first_name} ({user_id})"
        forum_topic = await context.bot.create_forum_topic(
            chat_id=chat_id,
            name=topic_name
        )
        topic_id = forum_topic.message_thread_id

        # Сохраняем соответствия
        topics[str(user_id)] = topic_id
        user_topics[str(topic_id)] = str(user_id)

        data["topics"] = topics
        data["user_topics"] = user_topics
        safe_json_write(data, DATA_FILE)

        return topic_id
    except telegram.error.TelegramError as e:
        print(f"Помилка створення теми: {e}")
        return None
    except Exception as e:
        print(f"Error in get_or_create_topic: {e}")
        return None


async def set_default_commands(application):
    try:
        commands = [
            BotCommand("start", "Запустити бота"),
            BotCommand("rate", "Залишити відгук"),
            BotCommand("message", "Почати введення повідомлень адміністраторам"),
            BotCommand("stopmessage", "Завершити введення повідомлень"),
            BotCommand("fromus", "Інформація про створювача"),
            BotCommand("help", "Показати доступні команди"),
        ]
        await application.bot.set_my_commands(commands, scope=BotCommandScopeDefault())
    except Exception as e:
        print(f"Error in set_default_commands: {e}")


async def set_creator_commands(application):
    try:
        commands = [
            BotCommand("mutelist", "Показати список замучених користувачів"),
            BotCommand("mute", "Замутити користувача"),
            BotCommand("unmute", "Розмутити користувача"),
            BotCommand("ban", "Забанити користувача"),
            BotCommand("unban", "Розбанити користувача"),
            BotCommand("alllist", "Показати всіх користувачів"),
            BotCommand("fromus", "Інформація про створювача"),
            BotCommand("help", "Показати доступні команди"),
            BotCommand("info", "Показати інформацію про програмістів та адміністраторів"),
            BotCommand("get_alllist", "Отримати Excel файл з користувачами"),
            BotCommand("set_alllist", "Записати Excel файл з користувачами"),
        ]
        await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CREATOR_CHAT_ID))
    except Exception as e:
        print(f"Error in set_creator_commands: {e}")

async def set_save_commands(application):
    commands = [
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
        BotCommand("get_logs", "Отримати логи"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=-1002648725095))


async def main():
    try:
        application = Application.builder().token(BOTTOCEN).build()

        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("rate", rate))
        application.add_handler(CommandHandler("message", message))
        application.add_handler(CommandHandler("stopmessage", stopmessage))
        application.add_handler(CommandHandler("fromus", fromus))
        application.add_handler(CommandHandler("help", help))
        application.add_handler(CommandHandler("mute", mute))
        application.add_handler(CommandHandler("unmute", unmute))
        application.add_handler(CommandHandler("ban", ban))
        application.add_handler(CommandHandler("unban", unban))
        application.add_handler(CommandHandler("mutelist", mutelist))
        application.add_handler(CommandHandler("alllist", alllist))
        application.add_handler(CommandHandler("allmessage", allmessage))
        application.add_handler(CommandHandler("admin", admin))
        application.add_handler(CommandHandler("deleteadmin", deleteadmin))
        application.add_handler(CommandHandler("programier", programier))
        application.add_handler(CommandHandler("deleteprogramier", deleteprogramier))
        application.add_handler(CommandHandler("info", info))
        application.add_handler(CommandHandler("get_alllist", get_alllist))
        application.add_handler(CommandHandler("set_alllist", set_alllist))
        application.add_handler(CommandHandler("get_logs", get_logs))

        application.add_handler(CallbackQueryHandler(button_callback))
        #application.add_handler(CallbackQueryHandler(button))
        application.add_handler(MessageHandler(filters.ALL, handle_message))

        await set_default_commands(application)
        await set_creator_commands(application)
        await set_save_commands(application)

        scheduler = AsyncIOScheduler(timezone=pytz.timezone("Europe/Kyiv"))
        scheduler.add_job(send_user_list, "cron", hour=0, minute=0)
        scheduler.add_job(check_mute_expirations, "interval", minutes=1)
        scheduler.start()

        application.run_polling()
    except Exception as e:
        print(f"Error in main: {e}")


if __name__ == "__main__":
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()
    asyncio.run(main())